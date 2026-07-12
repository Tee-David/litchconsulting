"""Golden-fixture suite for the Cash Flow Statement (Direct method) compiler."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.cashflow import COMPILER_VERSION, compile_cashflow
from litchai.contracts.cashflow import CashflowContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["cashflow_basic.json", "cashflow_service.json"]


def load_fixture(name: str) -> CashflowContract:
    return CashflowContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: CashflowContract) -> dict[str, float]:
    net_op = sum(i.amount for i in c.operating_receipts) - sum(i.amount for i in c.operating_payments)
    net_inv = sum(i.amount for i in c.investing_receipts) - sum(i.amount for i in c.investing_payments)
    net_fin = sum(i.amount for i in c.financing_receipts) - sum(i.amount for i in c.financing_payments)
    net_change = net_op + net_inv + net_fin
    return {
        "net_operating": net_op,
        "net_investing": net_inv,
        "net_financing": net_fin,
        "net_change": net_change,
        "closing_cash": c.opening_cash + net_change,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_cashflow(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_cashflow(load_fixture("cashflow_basic.json")).compiler_version == COMPILER_VERSION


def test_key_cells_are_formulas(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["Cash Flow"]
    for name in ("net_operating", "net_investing", "net_financing", "net_change", "closing_cash"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_no_cached_results(compiled):
    _, result, path = compiled
    ws = load_workbook(path, data_only=True)["Cash Flow"]
    for name, ref in result.key_cells.items():
        assert ws[ref].value is None, f"{name} at {ref} has a cached value"


def test_recompute_matches_golden(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_closing_ties_to_opening_plus_change(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    closing = recompute.value_at(grid, result.key_cells["closing_cash"])
    net_change = recompute.value_at(grid, result.key_cells["net_change"])
    assert closing == pytest.approx(contract.opening_cash + net_change, abs=0.01)
