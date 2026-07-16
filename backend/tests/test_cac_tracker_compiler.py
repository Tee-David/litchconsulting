"""Golden-fixture suite for the CAC Annual Return & Compliance Tracker (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.cac_tracker import COMPILER_VERSION, compile_cac_tracker
from litchai.contracts.cac_tracker import CacTrackerContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"


def load_fixture(name: str) -> CacTrackerContract:
    return CacTrackerContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: CacTrackerContract) -> dict[str, float]:
    total = len(c.items)
    completed = sum(1 for i in c.items if i.completed)
    overdue = sum(1 for i in c.items if not i.completed and c.as_at > i.due_date)
    return {
        "total_items": total,
        "completed_count": completed,
        "outstanding_count": total - completed,
        "overdue_count": overdue,
        "completion_pct": completed / total * 100,
        "total_fees": sum(i.fee for i in c.items),
        "outstanding_fees": sum(i.fee for i in c.items if not i.completed),
    }


@pytest.fixture()
def compiled(tmp_path):
    contract = load_fixture("cac_tracker_basic.json")
    result = compile_cac_tracker(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_cac_tracker(load_fixture("cac_tracker_basic.json")).compiler_version == COMPILER_VERSION


def test_summary_cells_are_formulas(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["Compliance"]
    for name in result.key_cells:
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_overdue_and_completion(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.value_at(grid, result.key_cells["overdue_count"]) == pytest.approx(2.0)
    assert recompute.value_at(grid, result.key_cells["completion_pct"]) == pytest.approx(40.0)
