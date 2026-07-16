"""Golden-fixture suite for the Statement of Affairs / Simple Balance Sheet (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.statement_of_affairs import COMPILER_VERSION, compile_statement_of_affairs
from litchai.contracts.statement_of_affairs import StatementOfAffairsContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["statement_of_affairs_balanced.json", "statement_of_affairs_unbalanced.json"]


def load_fixture(name: str) -> StatementOfAffairsContract:
    return StatementOfAffairsContract.model_validate(
        json.loads((FIXTURES / name).read_text(encoding="utf-8"))
    )


def golden(c: StatementOfAffairsContract) -> dict[str, float]:
    total_assets = sum(i.amount for i in (*c.non_current_assets, *c.current_assets))
    total_liab = sum(i.amount for i in (*c.current_liabilities, *c.non_current_liabilities))
    net_assets = total_assets - total_liab
    total_equity = sum(i.amount for i in c.equity)
    return {
        "total_assets": total_assets,
        "total_liabilities": total_liab,
        "net_assets": net_assets,
        "total_equity": total_equity,
        "balance_check": net_assets - total_equity,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_statement_of_affairs(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return request.param, contract, result, path


def test_compiler_version_recorded():
    got = compile_statement_of_affairs(load_fixture(FIXTURE_NAMES[0]))
    assert got.compiler_version == COMPILER_VERSION


def test_computed_cells_are_formulas(compiled):
    _, _, result, path = compiled
    ws = load_workbook(path)["Statement of Affairs"]
    for name in ("total_assets", "total_liabilities", "net_assets", "total_equity", "balance_check"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    _, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_balance_status(compiled):
    name, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    check = recompute.value_at(grid, result.key_cells["balance_check"])
    if name == "statement_of_affairs_balanced.json":
        assert check == pytest.approx(0.0, abs=0.01)
    else:
        assert abs(check) > 0.01
        assert check == pytest.approx(golden(contract)["balance_check"], abs=0.01)
