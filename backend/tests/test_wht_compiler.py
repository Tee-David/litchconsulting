"""Golden-fixture suite for the WHT Schedule & Credit Note Reconciliation (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.wht import COMPILER_VERSION, compile_wht_schedule
from litchai.contracts.wht import WhtScheduleContract
from litchai.taxconfig import load_tax_config
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["wht_schedule_balanced.json", "wht_schedule_gap.json"]
CFG = load_tax_config()


def load_fixture(name: str) -> WhtScheduleContract:
    return WhtScheduleContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def _rate(category: str, payee_type: str) -> float:
    return CFG["wht"]["rates"][category][f"{payee_type}Pct"] / 100.0


def golden(c: WhtScheduleContract) -> dict[str, float]:
    deducted = sum(
        0.0 if line.exempt else line.gross_amount * _rate(line.category, line.payee_type)
        for line in c.deductions
    )
    credits = sum(n.amount for n in c.credit_notes)
    return {
        "total_wht_deducted": deducted,
        "total_credit_notes": credits,
        "difference": deducted - credits,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_wht_schedule(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return request.param, contract, result, path


def test_compiler_version_recorded():
    assert compile_wht_schedule(load_fixture("wht_schedule_balanced.json")).compiler_version == COMPILER_VERSION


def test_wht_and_totals_are_formulas(compiled):
    _, _, result, path = compiled
    ws = load_workbook(path)["WHT Schedule"]
    for name in ("total_wht_deducted", "total_credit_notes", "difference"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    _, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_reconciliation_status(compiled):
    """Balanced fixture ties (difference 0); the gap fixture surfaces the missing note."""
    name, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    diff = recompute.value_at(grid, result.key_cells["difference"])
    if name == "wht_schedule_balanced.json":
        assert diff == pytest.approx(0.0, abs=0.01)
    else:
        assert diff > 0.01  # a credit note is missing
        assert diff == pytest.approx(golden(contract)["difference"], abs=0.01)
