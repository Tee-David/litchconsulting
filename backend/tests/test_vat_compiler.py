"""Golden-fixture suite for the VAT Returns Pack compiler (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.vat import COMPILER_VERSION, compile_vat_return
from litchai.contracts.vat import VatReturnContract
from litchai.taxconfig import load_tax_config
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["vat_return_basic.json", "vat_return_credit.json"]
RATE = load_tax_config()["vat"]["standardRatePct"] / 100.0


def load_fixture(name: str) -> VatReturnContract:
    return VatReturnContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: VatReturnContract) -> dict[str, float]:
    output_vat = sum(i.net_amount for i in c.standard_rated_sales) * RATE
    input_vat = sum(i.net_amount for i in c.standard_rated_purchases) * RATE
    total_sales_net = sum(
        i.net_amount
        for i in (*c.standard_rated_sales, *c.zero_rated_sales, *c.exempt_sales)
    )
    return {
        "output_vat": output_vat,
        "input_vat": input_vat,
        "total_sales_net": total_sales_net,
        "net_vat": output_vat - input_vat - c.vat_credit_brought_forward,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_vat_return(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return request.param, contract, result, path


def test_compiler_version_recorded():
    assert compile_vat_return(load_fixture("vat_return_basic.json")).compiler_version == COMPILER_VERSION


def test_computed_cells_are_formulas_not_cached_values(compiled):
    _, _, result, path = compiled
    ws = load_workbook(path)["VAT Return"]
    for name in ("output_vat", "input_vat", "net_vat", "total_sales_net"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_rate_comes_from_config_as_an_input_cell(compiled):
    _, _, _, path = compiled
    ws = load_workbook(path)["VAT Return"]
    assert ws["C6"].value == pytest.approx(RATE)  # 0.075, a literal input — not hardcoded VAT


def test_recompute_matches_golden(compiled):
    _, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_credit_position_is_negative(compiled):
    """When input VAT exceeds output + credit b/f, net VAT is a credit c/f (<0)."""
    name, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    net = recompute.value_at(grid, result.key_cells["net_vat"])
    if name == "vat_return_credit.json":
        assert net < 0
    assert net == pytest.approx(golden(contract)["net_vat"], abs=0.01)
