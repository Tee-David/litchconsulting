"""Golden-fixture suite for the AR/AP Aging compiler (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.ar_ap_aging import COMPILER_VERSION, compile_ar_ap_aging
from litchai.contracts.ar_ap_aging import AgingContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
BUCKETS = ["not_due", "b1_30", "b31_60", "b61_90", "b90_plus"]


def load_fixture(name: str) -> AgingContract:
    return AgingContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def _bucket(days: int) -> str:
    if days <= 0:
        return "not_due"
    if days <= 30:
        return "b1_30"
    if days <= 60:
        return "b31_60"
    if days <= 90:
        return "b61_90"
    return "b90_plus"


def golden(c: AgingContract) -> dict[str, float]:
    totals = {b: 0.0 for b in BUCKETS}
    for item in c.items:
        totals[_bucket((c.as_at - item.due_date).days)] += item.amount
    return {
        "grand_total": sum(i.amount for i in c.items),
        "bucket_check": 0.0,
        **{f"total_{b}": v for b, v in totals.items()},
    }


@pytest.fixture()
def compiled(tmp_path):
    contract = load_fixture("ar_ap_aging_debtors.json")
    result = compile_ar_ap_aging(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_ar_ap_aging(load_fixture("ar_ap_aging_debtors.json")).compiler_version == COMPILER_VERSION


def test_bucket_totals_are_formulas(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["Aging"]
    for name in ("grand_total", "bucket_check", *[f"total_{b}" for b in BUCKETS]):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_one_item_lands_in_each_bucket(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    for b in BUCKETS:
        assert recompute.value_at(grid, result.key_cells[f"total_{b}"]) > 0, b
    assert recompute.value_at(grid, result.key_cells["bucket_check"]) == pytest.approx(0.0, abs=0.01)
