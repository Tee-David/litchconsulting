"""Golden-fixture suite for the IAS 1 annual-report compiler.

Layers, mirroring the single-sheet suites: structural (formulas only,
cross-sheet links, no cached values, injection guard), recompute (zero error
tokens on every sheet, every *_check cell exactly zero, golden mirrors match
for both year columns), and behavioural edges (unbalanced exact gap, minimal
contract with emptied sections, deterministic bytes).
"""
import hashlib

import pytest
from openpyxl import load_workbook

from annual_report_helpers import fixture_annual_report, golden_ias1, load_annual_fixture
from litchai.compilers.annual_report import (
    COMPILER_VERSION,
    SHEETS_IAS1,
    compile_annual_report,
)
from litchai.compilers.annual_report._rows import save_workbook_deterministic
from litchai.validation import recompute

FIXTURE = "annual_report_ias1_basic.json"


@pytest.fixture(scope="module")
def compiled(tmp_path_factory):
    contract = load_annual_fixture(FIXTURE)
    result = compile_annual_report(contract)
    path = tmp_path_factory.mktemp("ias1") / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


@pytest.fixture(scope="module")
def grids(compiled):
    _, _, path = compiled
    return recompute.recompute_workbook(path)


def test_compiler_version_and_tab_order(compiled):
    _, result, path = compiled
    assert result.compiler_version == COMPILER_VERSION
    assert load_workbook(path).sheetnames == list(SHEETS_IAS1)


def test_every_key_cell_is_a_formula(compiled):
    _, result, path = compiled
    wb = load_workbook(path)
    for name, qualified in result.key_cells.items():
        sheet, ref = recompute.split_sheet_ref(qualified)
        value = wb[sheet][ref].value
        assert isinstance(value, str) and value.startswith("="), (
            f"{name} at {qualified} is not a formula: {value!r}"
        )


def test_cross_sheet_links_present(compiled):
    _, _, path = compiled
    wb = load_workbook(path)

    def formulas(sheet):
        return [
            c.value
            for row in wb[sheet].iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        ]

    assert any("Schedules!" in f for f in formulas("P&L"))
    assert any("'Bank Recon'!" in f for f in formulas("SOFP"))
    assert any("'P&L'!" in f for f in formulas("SOCF"))


def test_no_cached_results_written(compiled):
    _, result, path = compiled
    wb = load_workbook(path, data_only=True)
    for name, qualified in result.key_cells.items():
        sheet, ref = recompute.split_sheet_ref(qualified)
        assert wb[sheet][ref].value is None, f"{name} at {qualified} has a cached value"


def test_label_injection_guard(tmp_path):
    hostile = "=HYPERLINK(\"http://evil.example\",\"click\")"
    contract = fixture_annual_report(
        schedules={"revenue": [{"label": hostile, "current": 227000, "prior": 191500}]}
    )
    path = tmp_path / "hostile.xlsx"
    compile_annual_report(contract).workbook.save(path)
    ws = load_workbook(path)["Schedules"]
    hit = [c for row in ws.iter_rows() for c in row if c.value == hostile]
    assert hit, "hostile label not found"
    assert all(c.data_type == "s" for c in hit), "label became a live formula"


def test_recompute_no_errors_all_sheets(grids):
    assert recompute.find_workbook_errors(grids) == []


def test_all_check_cells_are_zero(compiled, grids):
    _, result, _ = compiled
    checks = [n for n in result.key_cells if n.split(":")[1].endswith("_check")]
    assert checks, "no check cells registered"
    for name in checks:
        got = recompute.value_at_ref(grids, result.key_cells[name])
        assert got == pytest.approx(0, abs=0.005), f"{name} recomputed to {got}"


def test_recompute_matches_golden_both_years(compiled, grids):
    contract, result, _ = compiled
    for name, expected in golden_ias1(contract).items():
        got = recompute.value_at_ref(grids, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_unbalanced_fixture_recomputes_exact_gap(tmp_path):
    contract = fixture_annual_report(
        sofp={"retained_earnings": {"current": 68050, "prior": 62650}}
    )
    result = compile_annual_report(contract)
    path = tmp_path / "unbalanced.xlsx"
    result.workbook.save(path)
    grids = recompute.recompute_workbook(path)
    assert recompute.value_at_ref(grids, result.key_cells["sofp:balance_check"]) == -500
    assert recompute.value_at_ref(grids, result.key_cells["sofp:balance_check:py"]) == 0


def test_minimal_contract_compiles_without_errors(tmp_path):
    # Structure only: emptied optional sections must not emit reversed SUM
    # ranges or dangling refs (balance checks are expected to be non-zero
    # because the trial balance no longer ties).
    contract = fixture_annual_report(
        schedules={
            "distribution_costs": [],
            "inventories": [],
            "ppe_classes": [
                {
                    "label": "Plant",
                    "cost_opening": 1000,
                    "dep_opening": 200,
                }
            ],
        },
        bank_recon={"unrecorded_debits": []},
    )
    result = compile_annual_report(contract)
    path = tmp_path / "minimal.xlsx"
    result.workbook.save(path)
    grids = recompute.recompute_workbook(path)
    assert recompute.find_workbook_errors(grids) == []
    assert recompute.value_at_ref(grids, result.key_cells["schedules:distribution_total"]) == 0
    assert recompute.value_at_ref(grids, result.key_cells["schedules:inventories_total"]) == 0
    assert recompute.value_at_ref(grids, result.key_cells["schedules:ppe_nbv_total"]) == 800


def test_identical_contract_produces_identical_bytes(tmp_path):
    contract = load_annual_fixture(FIXTURE)
    digests = []
    for i in range(2):
        path = tmp_path / f"det{i}.xlsx"
        save_workbook_deterministic(compile_annual_report(contract).workbook, path)
        digests.append(hashlib.sha256(path.read_bytes()).hexdigest())
    assert digests[0] == digests[1]
