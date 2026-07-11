"""Golden-fixture suite for the P&L compiler (PRD G1: this gate is 100%-pass
before any file may reach HITL review).

Two layers:
1. Structural — every computed cell is a formula, never a value (PRD §3).
2. Recompute — headless LibreOffice evaluates the workbook; zero error tokens,
   and every named total matches the golden expectation (FR6).
"""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.pnl import COMPILER_VERSION, compile_pnl
from litchai.contracts.pnl import PnLContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["pnl_basic.json", "pnl_messy.json", "pnl_service.json"]


def load_fixture(name: str) -> PnLContract:
    data = json.loads((FIXTURES / name).read_text(encoding="utf-8"))
    return PnLContract.model_validate(data)


def golden_totals(c: PnLContract) -> dict[str, float]:
    rev = sum(i.amount for i in c.revenue)
    cogs = sum(i.amount for i in c.cost_of_sales)
    opex = sum(i.amount for i in c.operating_expenses)
    other = sum(i.amount for i in c.other_income)
    return {
        "total_revenue": rev,
        "gross_profit": rev - cogs,
        "operating_profit": rev - cogs - opex,
        "net_profit": rev - cogs - opex + other,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_pnl(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_pnl(load_fixture("pnl_basic.json")).compiler_version == COMPILER_VERSION


def test_computed_cells_are_formulas(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["P&L"]
    for name in ("total_revenue", "gross_profit", "operating_profit", "net_profit"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), (
            f"{name} at {result.key_cells[name]} is not a formula: {value!r}"
        )


def test_no_cached_results_written(compiled):
    # data_only=True exposes cached formula results; a compiler-fresh file must
    # have none — proof that Python arithmetic never entered the sheet.
    _, result, path = compiled
    ws = load_workbook(path, data_only=True)["P&L"]
    for name, ref in result.key_cells.items():
        assert ws[ref].value is None, f"{name} at {ref} has a cached value"


def test_libreoffice_recompute_matches_golden(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden_totals(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name
