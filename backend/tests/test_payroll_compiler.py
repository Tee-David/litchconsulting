"""Golden-fixture suite for the PAYE Payroll Run compiler.

The Python PAYE reference below mirrors frontend/src/lib/calculators/paye.ts
(chargeable = gross − pension − NHF, then progressive bands) and reads the same
shared config the compiler embeds into its Excel formula. The recompute test
asserting LibreOffice == this reference is what proves the generated formula and
the law agree.
"""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.payroll import COMPILER_VERSION, compile_payroll
from litchai.contracts.payroll import Employee, PayrollContract
from litchai.taxconfig import load_tax_config
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["payroll_basic.json", "payroll_single.json"]
CFG = load_tax_config()


def load_fixture(name: str) -> PayrollContract:
    return PayrollContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def paye_on(chargeable: float) -> float:
    """Progressive PAYE reference — mirrors taxOnChargeable() in paye.ts."""
    tax = 0.0
    remaining = max(0.0, chargeable)
    for band in CFG["paye"]["bands"]:
        if remaining <= 0:
            break
        width = band["width"] if band["width"] is not None else remaining
        taxable = min(remaining, width)
        tax += taxable * (band["ratePct"] / 100)
        remaining -= taxable
    return tax


def expected_employee(emp: Employee) -> dict[str, float]:
    pension = emp.gross_annual * (CFG["payroll"]["pension"]["employeePct"] / 100) if emp.pension else 0
    nhf = emp.gross_annual * (CFG["payroll"]["nhf"]["employeePct"] / 100) if emp.nhf else 0
    chargeable = emp.gross_annual - pension - nhf
    paye = paye_on(chargeable)
    return {"paye": paye, "net": emp.gross_annual - pension - nhf - paye}


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_payroll(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_payroll(load_fixture("payroll_single.json")).compiler_version == COMPILER_VERSION


def test_key_cells_are_formulas(compiled):
    contract, result, path = compiled
    ws = load_workbook(path)["Payroll"]
    for i in range(len(contract.employees)):
        for name in (f"paye_{i}", f"net_{i}"):
            value = ws[result.key_cells[name]].value
            assert isinstance(value, str) and value.startswith("="), name
    for name in ("total_gross", "total_paye", "total_net"):
        assert str(ws[result.key_cells[name]].value).startswith("=")


def test_no_cached_results(compiled):
    _, result, path = compiled
    ws = load_workbook(path, data_only=True)["Payroll"]
    for name, ref in result.key_cells.items():
        assert ws[ref].value is None, f"{name} at {ref} has a cached value"


def test_recompute_paye_matches_reference(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []

    total_paye = 0.0
    total_net = 0.0
    total_gross = 0.0
    for i, emp in enumerate(contract.employees):
        exp = expected_employee(emp)
        got_paye = recompute.value_at(grid, result.key_cells[f"paye_{i}"])
        got_net = recompute.value_at(grid, result.key_cells[f"net_{i}"])
        assert got_paye == pytest.approx(exp["paye"], abs=0.01), f"{emp.name} PAYE"
        assert got_net == pytest.approx(exp["net"], abs=0.01), f"{emp.name} net"
        total_paye += exp["paye"]
        total_net += exp["net"]
        total_gross += emp.gross_annual

    assert recompute.value_at(grid, result.key_cells["total_paye"]) == pytest.approx(total_paye, abs=0.01)
    assert recompute.value_at(grid, result.key_cells["total_net"]) == pytest.approx(total_net, abs=0.01)
    assert recompute.value_at(grid, result.key_cells["total_gross"]) == pytest.approx(total_gross, abs=0.01)
