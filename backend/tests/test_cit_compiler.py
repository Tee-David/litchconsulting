"""Golden-fixture suite for the CIT Computation & Capital Allowance Register (v1.1)."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.cit import COMPILER_VERSION, compile_cit
from litchai.contracts.cit import CitContract
from litchai.taxconfig import load_tax_config
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["cit_standard.json", "cit_small.json"]
CIT = load_tax_config()["cit"]


def load_fixture(name: str) -> CitContract:
    return CitContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: CitContract) -> dict[str, float]:
    addbacks = sum(a.amount for a in c.disallowable_addbacks)
    deductions = sum(a.amount for a in c.other_deductions)
    total_ca = sum(a.cost * a.allowance_rate_pct / 100.0 for a in c.capital_allowance_assets)
    assessable = c.net_profit_per_accounts + addbacks - total_ca - deductions
    is_small = int(
        c.turnover <= CIT["smallCompany"]["maxTurnover"]
        and c.total_fixed_assets <= CIT["smallCompany"]["maxFixedAssets"]
        and not c.professional_services
    )
    rate = (CIT["smallCompany"]["ratePct"] if is_small else CIT["standardRatePct"]) / 100.0
    cit_payable = assessable * rate
    levy = 0.0 if is_small else assessable * CIT["developmentLevy"]["ratePct"] / 100.0
    return {
        "total_capital_allowances": total_ca,
        "is_small": float(is_small),
        "assessable_profit": assessable,
        "cit_payable": cit_payable,
        "development_levy": levy,
        "total_tax": cit_payable + levy,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_cit(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return request.param, contract, result, path


def test_compiler_version_recorded():
    assert compile_cit(load_fixture("cit_standard.json")).compiler_version == COMPILER_VERSION


def test_computed_cells_are_formulas(compiled):
    _, _, result, path = compiled
    ws = load_workbook(path)["CIT Computation"]
    for name in ("assessable_profit", "cit_payable", "development_levy", "total_tax", "is_small"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    _, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_small_company_pays_zero(compiled):
    name, _, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    total = recompute.value_at(grid, result.key_cells["total_tax"])
    is_small = recompute.value_at(grid, result.key_cells["is_small"])
    if name == "cit_small.json":
        assert is_small == pytest.approx(1.0)
        assert total == pytest.approx(0.0, abs=0.01)   # 0% CIT + levy-exempt
    else:
        assert is_small == pytest.approx(0.0)
        assert total > 0
