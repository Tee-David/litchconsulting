"""Golden-fixture suite for the IFRS 18 annual-report variant.

Beyond the IAS 1 layers: the two mandatory subtotals have the template's
formula shapes, the cash-flow statement starts from operating profit (not
profit before tax), the MPM note reconciles (and nil-discloses when empty),
and the shared sheets are cell-identical across the two variants.
"""
import re

import pytest
from openpyxl import load_workbook

from annual_report_helpers import fixture_annual_report, golden_ifrs18, load_annual_fixture
from litchai.compilers.annual_report import SHEETS_IFRS18, compile_annual_report
from litchai.validation import recompute

FIXTURE = "annual_report_ifrs18_basic.json"


@pytest.fixture(scope="module")
def compiled(tmp_path_factory):
    contract = load_annual_fixture(FIXTURE)
    result = compile_annual_report(contract)
    path = tmp_path_factory.mktemp("ifrs18") / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


@pytest.fixture(scope="module")
def grids(compiled):
    _, _, path = compiled
    return recompute.recompute_workbook(path)


def test_tab_order(compiled):
    _, _, path = compiled
    assert load_workbook(path).sheetnames == list(SHEETS_IFRS18)


def test_mandatory_subtotal_shapes(compiled):
    _, result, path = compiled
    wb = load_workbook(path)

    def formula_at(key):
        sheet, ref = recompute.split_sheet_ref(result.key_cells[key])
        return wb[sheet][ref].value

    # Operating profit: gross profit + the operating-category block.
    assert re.fullmatch(r"=D\d+\+SUM\(D\d+:D\d+\)", formula_at("pnl:operating_profit"))
    # PBFIT: operating + investing total + financing total.
    assert re.fullmatch(r"=D\d+\+D\d+\+D\d+", formula_at("pnl:pbfit"))


def test_socf_starts_from_operating_profit(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["SOCF (IFRS 18)"]
    start_formula = f"={result.key_cells['pnl:operating_profit']}"
    cells = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert start_formula in cells
    assert "pnl:profit_before_tax" not in result.key_cells  # IAS 1-only subtotal


def test_recompute_no_errors_all_sheets(grids):
    assert recompute.find_workbook_errors(grids) == []


def test_all_check_cells_are_zero(compiled, grids):
    _, result, _ = compiled
    checks = [n for n in result.key_cells if n.split(":")[1].endswith("_check")]
    for name in checks:
        got = recompute.value_at_ref(grids, result.key_cells[name])
        assert got == pytest.approx(0, abs=0.005), f"{name} recomputed to {got}"


def test_recompute_matches_golden_both_years(compiled, grids):
    contract, result, _ = compiled
    for name, expected in golden_ifrs18(contract).items():
        got = recompute.value_at_ref(grids, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_nil_disclosure_when_no_mpms(compiled):
    _, _, path = compiled
    ws = load_workbook(path)["MPM Disclosure Note"]
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(t.startswith("Nil disclosure:") for t in texts)


def test_mpm_reconciliation_recomputes(tmp_path):
    contract = fixture_annual_report(
        standard="ifrs18",
        mpms=[
            {
                "name": "Adjusted operating profit",
                "rationale": "Excludes one-off restructuring and impairment charges.",
                "comparable_subtotal_label": "Operating profit or loss",
                "comparable_subtotal": {"current": 10100, "prior": 9300},
                "reconciling_items": [
                    {"label": "Add back: restructuring costs", "current": 800, "prior": 0},
                    {"label": "Add back: impairment charge", "current": 200, "prior": 150},
                ],
                "tax_effect": {"current": -300, "prior": -45},
                "nci_effect": {"current": 0, "prior": 0},
            }
        ],
    )
    result = compile_annual_report(contract)
    path = tmp_path / "mpm.xlsx"
    result.workbook.save(path)
    grids = recompute.recompute_workbook(path)
    assert recompute.find_workbook_errors(grids) == []
    assert recompute.value_at_ref(grids, result.key_cells["mpm:mpm1_total"]) == pytest.approx(
        10100 + 800 + 200 - 300
    )
    assert recompute.value_at_ref(grids, result.key_cells["mpm:mpm1_total:py"]) == pytest.approx(
        9300 + 150 - 45
    )


def test_shared_sheets_identical_across_variants(tmp_path):
    ias1 = compile_annual_report(load_annual_fixture("annual_report_ias1_basic.json"))
    ifrs18 = compile_annual_report(load_annual_fixture(FIXTURE))
    p1, p2 = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    ias1.workbook.save(p1)
    ifrs18.workbook.save(p2)
    wb1, wb2 = load_workbook(p1), load_workbook(p2)
    for sheet in ("SOFP", "Bank Recon", "Schedules"):
        cells1 = {
            c.coordinate: c.value for row in wb1[sheet].iter_rows() for c in row
            if c.value is not None
        }
        cells2 = {
            c.coordinate: c.value for row in wb2[sheet].iter_rows() for c in row
            if c.value is not None
        }
        assert cells1 == cells2, f"{sheet} differs between variants"
