"""Multi-sheet recompute foundation for the annual-report compilers.

The CSV path (recompute_to_grid) exports only the first sheet, so a
multi-sheet workbook would gate vacuously. These tests prove the xlsx
round-trip evaluates cross-sheet formulas on every sheet, keeps sheet names
(including quoting-hostile ones) intact, and that sheet-qualified refs
resolve through value_at_ref.
"""
import pytest
from openpyxl import Workbook

from litchai.compilers._common import sheet_ref, split_sheet_ref
from litchai.validation import recompute


def test_sheet_ref_quoting():
    assert sheet_ref("SOFP", "D26") == "SOFP!D26"
    assert sheet_ref("Schedules", "D12") == "Schedules!D12"
    assert sheet_ref("Bank Recon", "F55") == "'Bank Recon'!F55"
    assert sheet_ref("P&L (IFRS 18)", "C5") == "'P&L (IFRS 18)'!C5"
    # A sheet named like a cell ref must be quoted even though it's plain.
    assert sheet_ref("A1", "B2") == "'A1'!B2"
    # Embedded apostrophes double, per Excel's escaping rule.
    assert sheet_ref("Tee's", "B2") == "'Tee''s'!B2"


def test_split_sheet_ref_round_trips():
    for sheet in ("SOFP", "Bank Recon", "P&L (IFRS 18)", "A1", "Tee's"):
        assert split_sheet_ref(sheet_ref(sheet, "D26")) == (sheet, "D26")
    assert split_sheet_ref("C5") == (None, "C5")


def build_linked_workbook(path, *, broken: bool = False):
    """Three sheets exercising both quoting cases; optionally a dangling
    cross-sheet ref that must surface as an error token after recompute."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["B2"] = 1200
    ws["B3"] = 800
    pnl = wb.create_sheet("P&L (IFRS 18)")
    pnl["C5"] = "=Inputs!B2+Inputs!B3"
    recon = wb.create_sheet("Bank Recon")
    recon["F55"] = "='P&L (IFRS 18)'!C5*2"
    if broken:
        recon["F60"] = "=Missing!A1"
    wb.save(path)
    return path


def test_recompute_workbook_evaluates_every_sheet(tmp_path):
    path = build_linked_workbook(tmp_path / "linked.xlsx")
    grids = recompute.recompute_workbook(path)
    assert set(grids) == {"Inputs", "P&L (IFRS 18)", "Bank Recon"}
    assert recompute.find_workbook_errors(grids) == []
    assert recompute.value_at_ref(grids, "Inputs!B2") == 1200
    assert recompute.value_at_ref(grids, "'P&L (IFRS 18)'!C5") == 2000
    assert recompute.value_at_ref(grids, "'Bank Recon'!F55") == 4000


def test_dangling_cross_sheet_ref_is_caught(tmp_path):
    path = build_linked_workbook(tmp_path / "broken.xlsx", broken=True)
    grids = recompute.recompute_workbook(path)
    errors = recompute.find_workbook_errors(grids)
    assert errors, "a formula against a missing sheet must yield an error token"
    assert all(sheet == "Bank Recon" for sheet, *_ in errors)


def test_value_at_ref_requires_known_qualified_sheet(tmp_path):
    path = build_linked_workbook(tmp_path / "linked.xlsx")
    grids = recompute.recompute_workbook(path)
    with pytest.raises(recompute.RecomputeError):
        recompute.value_at_ref(grids, "C5")  # unqualified
    with pytest.raises(recompute.RecomputeError):
        recompute.value_at_ref(grids, "Nowhere!C5")
