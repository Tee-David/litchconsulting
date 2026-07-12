"""Golden-fixture suite for the Bank Reconciliation compiler."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.bank_rec import COMPILER_VERSION, compile_bank_rec
from litchai.contracts.bank_rec import BankRecContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["bank_rec_balanced.json", "bank_rec_unbalanced.json"]


def load_fixture(name: str) -> BankRecContract:
    return BankRecContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: BankRecContract) -> dict[str, float]:
    adj_bank = (
        c.balance_per_bank
        + sum(i.amount for i in c.deposits_in_transit)
        - sum(i.amount for i in c.outstanding_cheques)
    )
    adj_book = (
        c.balance_per_books
        + sum(i.amount for i in c.add_to_books)
        - sum(i.amount for i in c.less_from_books)
    )
    return {
        "adjusted_bank": adj_bank,
        "adjusted_book": adj_book,
        "difference": adj_bank - adj_book,
    }


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_bank_rec(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return request.param, contract, result, path


def test_compiler_version_recorded():
    assert compile_bank_rec(load_fixture("bank_rec_balanced.json")).compiler_version == COMPILER_VERSION


def test_key_cells_are_formulas(compiled):
    _, _, result, path = compiled
    ws = load_workbook(path)["Bank Rec"]
    for name in ("adjusted_bank", "adjusted_book", "difference"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_recompute_matches_golden(compiled):
    _, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_reconciliation_status(compiled):
    """Balanced fixture ties out (difference 0); unbalanced surfaces the exact gap."""
    name, contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    diff = recompute.value_at(grid, result.key_cells["difference"])
    if name == "bank_rec_balanced.json":
        assert diff == pytest.approx(0.0, abs=0.01)
    else:
        assert abs(diff) > 0.01
        assert diff == pytest.approx(golden(contract)["difference"], abs=0.01)
