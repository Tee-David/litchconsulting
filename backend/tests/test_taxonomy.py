"""Taxonomy suite: loader validation, the template-label bijection (the two
firm workbooks are taxonomy fixtures), the template-drift fingerprint, and
the operator-facing xlsx round-trip.

The workbook-dependent tests skip when plans/ isn't present (it is
gitignored; these run on the dev machine, not the VM).
"""
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.taxonomy import SUSPENSE_CODE, TaxonomyError, load_taxonomy
from litchai.taxonomy.export import export_taxonomy_xlsx, import_taxonomy_xlsx

PLANS = Path(__file__).resolve().parents[2] / "plans"
TEMPLATES = {
    "IAS 1 SME Annual Reporting Template.xlsx": (
        "6ccef44aebd21c08bbee8c566365e8526d4fc43f2ee028d4b11dcb2228b9dd67"
    ),
    "IFRS 18 Annual Report Template.xlsx": (
        "4f59466df9f4fc8672cd193fcf204a81247eea2489ee26a4f0993dd6645fb292"
    ),
}

needs_templates = pytest.mark.skipif(
    not all((PLANS / name).exists() for name in TEMPLATES),
    reason="firm template workbooks not present (plans/ is local-only)",
)


def test_packaged_taxonomy_loads():
    taxonomy = load_taxonomy()
    assert taxonomy.version == "2026.07.0"
    leaves = {c.code for c in taxonomy.postable_leaves()}
    assert SUSPENSE_CODE in leaves
    assert "transfers.internal" in leaves
    assert "bank.charges" in leaves


def _write(tmp_path, data) -> Path:
    path = tmp_path / "tax.json"
    path.write_text(json.dumps(data), encoding="utf-8")
    return path


def _base(categories):
    suspense = [
        {"code": "suspense", "label": "Suspense"},
        {
            "code": "suspense.uncategorized",
            "label": "Unclassified",
            "parent": "suspense",
            "postable": True,
            "nature": "expense",
        },
    ]
    return {"version": "0.0.1", "categories": categories + suspense}


@pytest.mark.parametrize(
    "categories, message",
    [
        (
            [
                {"code": "a", "label": "A"},
                {"code": "a", "label": "A again"},
            ],
            "duplicate code",
        ),
        ([{"code": "a", "label": "A", "parent": "ghost"}], "unknown parent"),
        (
            [{"code": "a", "label": "A", "postable": True}],
            "lacks nature",
        ),
        (
            [
                {"code": "a", "label": "A", "postable": True, "nature": "income"},
                {"code": "a.b", "label": "B", "parent": "a"},
            ],
            "must be a leaf",
        ),
        (
            [
                {"code": "a", "label": "A", "template_labels": ["Same"]},
                {"code": "b", "label": "B", "template_labels": ["Same"]},
            ],
            "claimed by both",
        ),
    ],
)
def test_loader_rejects_invalid(tmp_path, categories, message):
    with pytest.raises(TaxonomyError, match=message):
        load_taxonomy(_write(tmp_path, _base(categories)))


def test_loader_requires_suspense(tmp_path):
    data = {
        "version": "0.0.1",
        "categories": [{"code": "a", "label": "A", "postable": True, "nature": "income"}],
    }
    with pytest.raises(TaxonomyError, match="escape hatch"):
        load_taxonomy(_write(tmp_path, data))


def _schedule_input_labels(path: Path) -> list[str]:
    """Column-B labels of schedule data rows: B is text and D is a literal
    number (totals/checks are formulas; the PPE grid is per-client asset
    classes, not categories)."""
    ws = load_workbook(path)["Schedules"]
    labels, in_ppe = [], False
    for row in ws.iter_rows(min_row=7):
        section = row[0].value
        if isinstance(section, str) and section.startswith("Schedule"):
            in_ppe = section.startswith("Schedule 5")
        if in_ppe:
            continue
        label, amount = row[1].value, row[3].value
        if isinstance(label, str) and isinstance(amount, (int, float)):
            labels.append(label)
    return labels


@needs_templates
@pytest.mark.parametrize("name", list(TEMPLATES))
def test_every_template_schedule_label_maps_to_one_category(name):
    index = load_taxonomy().template_label_index()
    labels = _schedule_input_labels(PLANS / name)
    assert labels, "no schedule lines extracted — extraction heuristic broke"
    unmapped = [label for label in labels if label not in index]
    assert not unmapped, f"template labels without a category: {unmapped}"


@needs_templates
@pytest.mark.parametrize("name, expected", TEMPLATES.items())
def test_template_fingerprint_unchanged(name, expected):
    # A firm edit to its templates must be a deliberate event: bump the
    # taxonomy/compiler versions, re-derive, then update this fingerprint.
    wb = load_workbook(PLANS / name)
    parts = [
        f"{ws.title}!{cell.coordinate}={cell.value!r}"
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    got = hashlib.sha256("\n".join(sorted(parts)).encode()).hexdigest()
    assert got == expected, f"{name} changed on disk — treat as a template revision"


def test_xlsx_export_import_round_trip(tmp_path):
    taxonomy = load_taxonomy()
    path = tmp_path / "taxonomy.xlsx"
    export_taxonomy_xlsx(taxonomy, path)
    round_tripped = import_taxonomy_xlsx(path)
    assert round_tripped.version == taxonomy.version
    assert round_tripped.derived_from == taxonomy.derived_from
    assert [c.model_dump() for c in round_tripped.categories] == [
        c.model_dump() for c in taxonomy.categories
    ]
