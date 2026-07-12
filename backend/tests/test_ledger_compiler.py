"""Golden-fixture suite for the General Ledger compiler."""
import json
from collections import defaultdict
from pathlib import Path

import pytest
from openpyxl import load_workbook

from litchai.compilers.ledger import COMPILER_VERSION, compile_ledger
from litchai.contracts.ledger import LedgerContract
from litchai.validation import recompute

FIXTURES = Path(__file__).parent.parent / "fixtures" / "synthetic"
FIXTURE_NAMES = ["ledger_basic.json", "ledger_messy.json"]


def load_fixture(name: str) -> LedgerContract:
    return LedgerContract.model_validate(json.loads((FIXTURES / name).read_text(encoding="utf-8")))


def golden(c: LedgerContract) -> dict[str, float]:
    total_in = sum(t.amount for t in c.transactions if t.direction == "in")
    total_out = sum(t.amount for t in c.transactions if t.direction == "out")
    return {"total_in": total_in, "total_out": total_out, "net_position": total_in - total_out}


def category_nets(c: LedgerContract) -> dict[str, float]:
    nets: dict[str, float] = defaultdict(float)
    for t in c.transactions:
        nets[t.category] += t.amount if t.direction == "in" else -t.amount
    return dict(nets)


@pytest.fixture(params=FIXTURE_NAMES)
def compiled(request, tmp_path):
    contract = load_fixture(request.param)
    result = compile_ledger(contract)
    path = tmp_path / "out.xlsx"
    result.workbook.save(path)
    return contract, result, path


def test_compiler_version_recorded():
    assert compile_ledger(load_fixture("ledger_basic.json")).compiler_version == COMPILER_VERSION


def test_key_cells_are_formulas(compiled):
    _, result, path = compiled
    ws = load_workbook(path)["Ledger"]
    for name in ("total_in", "total_out", "net_position"):
        value = ws[result.key_cells[name]].value
        assert isinstance(value, str) and value.startswith("="), name


def test_no_cached_results(compiled):
    _, result, path = compiled
    ws = load_workbook(path, data_only=True)["Ledger"]
    for name, ref in result.key_cells.items():
        assert ws[ref].value is None, f"{name} at {ref} has a cached value"


def test_recompute_matches_golden(compiled):
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    assert recompute.find_error_cells(grid) == []
    for name, expected in golden(contract).items():
        got = recompute.value_at(grid, result.key_cells[name])
        assert got == pytest.approx(expected, abs=0.01), name


def test_category_sumifs_tie_to_transactions(compiled):
    """Every per-category SUMIF net matches the direct Python total, and the
    category nets sum back to the overall net position."""
    contract, result, path = compiled
    grid = recompute.recompute_to_grid(path)
    ws = load_workbook(path)["Ledger"]
    expected_nets = category_nets(contract)

    # Locate the summary block by its "Category" header, then read each row.
    net_col = "G"
    seen: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        label = ws[f"C{r}"].value
        if label in expected_nets and isinstance(ws[f"{net_col}{r}"].value, str):
            seen[label] = recompute.value_at(grid, f"{net_col}{r}")
    assert set(seen) == set(expected_nets)
    for cat, net in expected_nets.items():
        assert seen[cat] == pytest.approx(net, abs=0.01), cat
    assert sum(seen.values()) == pytest.approx(
        golden(contract)["net_position"], abs=0.01
    )
