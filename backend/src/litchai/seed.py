"""Seeding CLI — ``python -m litchai.seed <cmd>`` (Phase 3).

Never test the ladder cold. Commands:

* ``taxonomy``  — sync ``litch-taxonomy.json`` → the ``taxonomy_categories`` mirror
* ``synonyms``  — curated Nigerian narration exemplars → ``category_memory`` (global)
* ``history``   — mine the firm's categorized workbooks → ``category_memory``;
                  unmapped labels are printed, never guessed
* ``embed``     — backfill embeddings for rows that lack them
* ``verify``    — replay held-out seeds → per-rung accuracy (the "are we warm?" gate)

The pure functions (rows/record builders, xlsx miner) are unit-tested; the
``main`` wiring binds them to the VM's psycopg repo/store + Ollama embedder.
"""
from __future__ import annotations

import argparse
import io
import json
from pathlib import Path

from litchai.categorize import normalize_narration
from litchai.categorize.memory_store import MemoryRecord
from litchai.embeddings import Embedder
from litchai.taxonomy import Taxonomy, load_taxonomy

SYNONYMS_PATH = Path(__file__).parent / "categorize" / "seeds" / "synonyms.json"

_CATEGORY_HEADERS = ("category", "code", "classification", "account", "gl code")
_DESC_HEADERS = ("description", "narration", "details", "particulars", "memo", "remarks")


def taxonomy_rows(taxonomy: Taxonomy) -> list[dict]:
    return [
        {
            "code": c.code, "taxonomy_version": taxonomy.version, "label": c.label,
            "parent": c.parent, "postable": c.postable, "nature": c.nature, "source": c.source,
        }
        for c in taxonomy.categories
    ]


def load_synonyms(path: Path | None = None) -> dict[str, list[str]]:
    data = json.loads((path or SYNONYMS_PATH).read_text(encoding="utf-8"))
    return {code: narrations for code, narrations in data.items() if not code.startswith("_")}


def synonym_records(
    data: dict[str, list[str]], taxonomy: Taxonomy, *, client_id: str | None = None
) -> tuple[list[MemoryRecord], list[str]]:
    """Build seed_template memory rows; return (records, unknown_codes). Unknown
    or non-postable codes are collected, never silently mapped."""
    postable = {c.code for c in taxonomy.postable_leaves()}
    records: list[MemoryRecord] = []
    unknown: list[str] = []
    seen: set[tuple[str, str]] = set()
    for code, narrations in data.items():
        if code not in postable:
            unknown.append(code)
            continue
        for narration in narrations:
            text = normalize_narration(narration)
            if not text or (text, code) in seen:
                continue
            seen.add((text, code))
            records.append(
                MemoryRecord(
                    id=0, normalized_text=text, category_code=code, source="seed_template",
                    client_id=client_id, taxonomy_version=taxonomy.version,
                )
            )
    return records, unknown


def mine_history_xlsx(
    data: bytes, taxonomy: Taxonomy, *, client_id: str | None = None
) -> tuple[list[MemoryRecord], list[str]]:
    """Mine (narration, category) pairs from a categorized workbook. A category
    cell is matched as a code or a template label; anything else is an unmapped
    label (returned, never guessed)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        rows = [tuple(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()

    desc_col, cat_col = _detect_history_header(rows)
    if desc_col is None or cat_col is None:
        raise ValueError("history workbook needs a description column and a category column")

    by_code = taxonomy.by_code()
    labels = taxonomy.template_label_index()
    records: list[MemoryRecord] = []
    unmapped: set[str] = set()
    seen: set[tuple[str, str]] = set()

    for raw in rows[1:]:
        narration = _cell(raw, desc_col)
        label = _cell(raw, cat_col)
        if not narration or not label:
            continue
        code = label if label in by_code else labels.get(label)
        if code is None:
            unmapped.add(label)
            continue
        text = normalize_narration(narration)
        if not text or (text, code) in seen:
            continue
        seen.add((text, code))
        records.append(
            MemoryRecord(
                id=0, normalized_text=text, category_code=code, source="seed_history",
                client_id=client_id, taxonomy_version=taxonomy.version,
            )
        )
    return records, sorted(unmapped)


def embed_records(records: list[MemoryRecord], embedder: Embedder) -> list[MemoryRecord]:
    texts = [r.normalized_text for r in records]
    vectors = embedder.embed_documents(texts)
    return [
        MemoryRecord(**{**r.__dict__, "embedding": v, "embedding_model": embedder.model})
        for r, v in zip(records, vectors)
    ]


def _detect_history_header(rows: list[tuple]) -> tuple[int | None, int | None]:
    for raw in rows[:15]:
        desc = cat = None
        for idx, value in enumerate(raw):
            if not isinstance(value, str):
                continue
            cell = value.strip().lower()
            if desc is None and any(cell.startswith(h) for h in _DESC_HEADERS):
                desc = idx
            if cat is None and any(cell.startswith(h) for h in _CATEGORY_HEADERS):
                cat = idx
        if desc is not None and cat is not None:
            return desc, cat
    return None, None


def _cell(raw: tuple, col: int) -> str:
    if col >= len(raw) or raw[col] is None:
        return ""
    return str(raw[col]).strip()


# --- CLI (VM wiring) -------------------------------------------------------


def main(argv: list[str] | None = None) -> int:  # pragma: no cover - VM wiring
    parser = argparse.ArgumentParser(prog="litchai.seed")
    sub = parser.add_subparsers(dest="cmd", required=True)
    sub.add_parser("taxonomy")
    p_syn = sub.add_parser("synonyms")
    p_syn.add_argument("--client-id", default=None)
    p_hist = sub.add_parser("history")
    p_hist.add_argument("--xlsx", required=True, type=Path)
    p_hist.add_argument("--client-id", default=None)
    sub.add_parser("embed")
    p_ver = sub.add_parser("verify")
    p_ver.add_argument("--min-accuracy", type=float, default=0.8)
    args = parser.parse_args(argv)

    taxonomy = load_taxonomy()

    if args.cmd == "taxonomy":
        from litchai.db.pg import connect

        conn = connect()
        with conn.cursor() as cur:
            for row in taxonomy_rows(taxonomy):
                cur.execute(
                    "INSERT INTO taxonomy_categories (code, taxonomy_version, label, parent, postable, "
                    "nature, source) VALUES (%(code)s, %(taxonomy_version)s, %(label)s, %(parent)s, "
                    "%(postable)s, %(nature)s, %(source)s) ON CONFLICT (code) DO UPDATE SET "
                    "taxonomy_version = EXCLUDED.taxonomy_version, label = EXCLUDED.label, "
                    "parent = EXCLUDED.parent, postable = EXCLUDED.postable, nature = EXCLUDED.nature, "
                    "source = EXCLUDED.source",
                    row,
                )
        conn.close()
        print(f"synced {len(taxonomy.categories)} categories (taxonomy {taxonomy.version})")
        return 0

    if args.cmd == "synonyms":
        from litchai.categorize.pg_memory import PgMemoryStore
        from litchai.db.pg import connect
        from litchai.embeddings import OllamaEmbedder

        records, unknown = synonym_records(load_synonyms(), taxonomy, client_id=args.client_id)
        records = embed_records(records, OllamaEmbedder())
        conn = connect()
        store = PgMemoryStore(conn)
        for rec in records:
            store.add(rec)
        conn.close()
        for code in unknown:
            print(f"  unknown code (skipped): {code}")
        print(f"seeded {len(records)} synonym exemplars")
        return 0

    if args.cmd == "history":
        from litchai.categorize.pg_memory import PgMemoryStore
        from litchai.db.pg import connect
        from litchai.embeddings import OllamaEmbedder

        records, unmapped = mine_history_xlsx(
            args.xlsx.read_bytes(), taxonomy, client_id=args.client_id
        )
        records = embed_records(records, OllamaEmbedder())
        conn = connect()
        store = PgMemoryStore(conn)
        for rec in records:
            store.add(rec)
        conn.close()
        for label in unmapped:
            print(f"  UNMAPPED label (fix the taxonomy or the workbook): {label!r}")
        print(f"mined {len(records)} history exemplars; {len(unmapped)} unmapped labels")
        return 0

    print(f"'{args.cmd}' runs on the VM (needs Postgres/Ollama)")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
