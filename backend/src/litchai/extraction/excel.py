"""Excel extraction engine (Phase 2b) — the zero-OCR-risk path, done first.

Fingerprints the workbook (``read_only`` + ``data_only``), finds the header row
heuristically, classifies the sheet, and emits rows with ``Sheet!cell``
provenance. Debit/credit columns net into a signed amount; a balance column, if
present, rides along for the balance-continuity gate.

Formula workbooks with no cached values (``data_only`` yields ``None``) are the
one case that needs LibreOffice — the caller recomputes once and re-runs; here we
simply flag the sheet ``uncached`` so the pipeline knows to do that pass.
"""
from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from litchai.extraction.base import ExtractedRow, ExtractionError, ExtractionResult, register_engine
from litchai.extraction.sandbox import check_zip_safety
from litchai.sanitize import parse_amount

_XLSX_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

# Header keyword → logical column role.
_HEADER_HINTS: dict[str, tuple[str, ...]] = {
    "date": ("date", "value date", "trans date", "txn date", "posting date"),
    "description": ("description", "narration", "details", "particulars", "remarks", "memo"),
    "debit": ("debit", "withdrawal", "withdrawals", "dr", "money out", "outflow"),
    "credit": ("credit", "deposit", "deposits", "lodgement", "cr", "money in", "inflow"),
    "amount": ("amount", "value"),
    "balance": ("balance", "running balance", "closing balance"),
}

_OPENING_LABELS = ("opening balance", "balance brought forward", "b/f", "bal b/f", "bbf")
_CLOSING_LABELS = ("closing balance", "balance carried forward", "c/f", "bal c/f", "bcf")


class ExcelEngine:
    name = "excel"

    def can_handle(self, mime: str, filename: str) -> bool:
        return mime in _XLSX_MIME or filename.lower().endswith((".xlsx", ".xlsm", ".xls"))

    def extract(self, data: bytes, *, page_range: tuple[int, int] | None = None) -> ExtractionResult:
        check_zip_safety(data)
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            sheet = _pick_sheet(wb)
            rows_raw = [tuple(r) for r in sheet.iter_rows(values_only=True)]
        finally:
            wb.close()

        header_idx, roles = _detect_header(rows_raw)
        if header_idx is None:
            raise ExtractionError("no recognizable header row (date/description/amount)")

        rows: list[ExtractedRow] = []
        opening = closing = None

        for r_offset, raw in enumerate(rows_raw[header_idx + 1 :], start=header_idx + 2):
            desc = _cell_text(raw, roles.get("description"))
            label = desc.lower().strip()

            balance = _amount_at(raw, roles.get("balance"))
            if any(k in label for k in _OPENING_LABELS):
                opening = balance if balance is not None else opening
                continue
            if any(k in label for k in _CLOSING_LABELS):
                closing = balance if balance is not None else closing
                continue

            amount, row_flags = _row_amount(raw, roles)
            if amount is None:
                continue  # not a transaction row (blank / subtotal without amount)

            desc_col = roles.get("description") or roles.get("amount") or 0
            rows.append(
                ExtractedRow(
                    raw_text=desc or "(no narration)",
                    amount=amount,
                    balance=balance,
                    sheet_ref=f"{sheet.title}!{get_column_letter(desc_col + 1)}{r_offset}",
                    row_ref=r_offset,
                    flags=tuple(row_flags),
                )
            )

        return ExtractionResult(
            engine=self.name,
            rows=rows,
            opening_balance=opening,
            closing_balance=closing,
            sheet_type=_classify(roles),
        )


def _pick_sheet(wb) -> object:
    """The sheet with the most non-empty rows — the statement, not a cover tab."""
    best, best_rows = None, -1
    for ws in wb.worksheets:
        n = ws.max_row or 0
        if n > best_rows:
            best, best_rows = ws, n
    if best is None:
        raise ExtractionError("empty workbook")
    return best


def _detect_header(rows_raw: list[tuple]) -> tuple[int | None, dict[str, int]]:
    for idx, raw in enumerate(rows_raw[:20]):
        roles: dict[str, int] = {}
        for c_idx, value in enumerate(raw):
            if not isinstance(value, str):
                continue
            cell = value.strip().lower()
            for role, hints in _HEADER_HINTS.items():
                if role in roles:
                    continue
                if any(cell == h or cell.startswith(h) for h in hints):
                    roles[role] = c_idx
        has_money = "amount" in roles or "debit" in roles or "credit" in roles
        if "description" in roles and has_money:
            return idx, roles
    return None, {}


def _classify(roles: dict[str, int]) -> str:
    if "balance" in roles and ("debit" in roles or "credit" in roles):
        return "bank_statement"
    if "debit" in roles or "credit" in roles or "amount" in roles:
        return "ledger"
    return "unknown"


def _cell_text(raw: tuple, col: int | None) -> str:
    if col is None or col >= len(raw) or raw[col] is None:
        return ""
    return str(raw[col]).strip()


def _amount_at(raw: tuple, col: int | None) -> Decimal | None:
    if col is None or col >= len(raw):
        return None
    parsed = parse_amount(raw[col])
    return parsed.value if parsed else None


def _row_amount(raw: tuple, roles: dict[str, int]) -> tuple[Decimal | None, list[str]]:
    """Signed amount for a row: credit − debit, or the single amount column."""
    flags: list[str] = []

    if "debit" in roles or "credit" in roles:
        debit = parse_amount(raw[roles["debit"]]) if "debit" in roles and roles["debit"] < len(raw) else None
        credit = parse_amount(raw[roles["credit"]]) if "credit" in roles and roles["credit"] < len(raw) else None
        if debit is None and credit is None:
            return None, flags
        value = (credit.value if credit else Decimal("0.00")) - (debit.value if debit else Decimal("0.00"))
        for part in (debit, credit):
            if part:
                flags.extend(part.flags)
        return value, flags

    if "amount" in roles:
        col = roles["amount"]
        parsed = parse_amount(raw[col]) if col < len(raw) else None
        if parsed is None:
            return None, flags
        return parsed.value, list(parsed.flags)

    return None, flags


def needs_recompute(data: bytes) -> bool:
    """True if the workbook has formulas but no cached values (``data_only``
    would yield ``None``). The pipeline then does one headless-LibreOffice
    recompute pass before extraction. A workbook of literal values → False.
    """
    check_zip_safety(data)
    formula = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    cached = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for f_ws, c_ws in zip(formula.worksheets, cached.worksheets):
            for f_row, c_row in zip(f_ws.iter_rows(), c_ws.iter_rows()):
                for f_cell, c_cell in zip(f_row, c_row):
                    if isinstance(f_cell.value, str) and f_cell.value.startswith("=") and c_cell.value is None:
                        return True
    finally:
        formula.close()
        cached.close()
    return False


register_engine(ExcelEngine())
