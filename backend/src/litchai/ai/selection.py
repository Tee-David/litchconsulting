"""AI-in-the-spreadsheet for the Analyses editor (Wave 2, Step 2).

The editor never ships the whole workbook to the model — only a *selection*: its
A1 range, the column headers, and the displayed cell values (plus any formulas).
An LLM proposes structured cell edits under a strict JSON schema; every proposed
FORMULA is then smoke-tested by writing the selection back into a throwaway
workbook at its real coordinates, recomputing it through LibreOffice and scanning
for error tokens — so a formula that would blow up (``#REF!``, ``#DIV/0!`` …)
comes back as a *warning*, not an applied edit.

The number is proposed by generation but **checked deterministically** — the same
contract the compile pipeline uses (see ``review/corrections.py``). Edits are only
ever *proposals*: the editor previews them and a human applies them.
"""
from __future__ import annotations

import logging

from litchai.ai.cache import AiCache, AiTelemetry
from litchai.ai.harness import run_task
from litchai.ai.provider import Provider
from litchai.ai.tasks import TaskPolicy, TaskSpec
from pydantic import BaseModel

log = logging.getLogger(__name__)

# The first six commands the editor exposes (Step 2c).
COMMANDS: tuple[str, ...] = (
    "explain",
    "write_formula",
    "clean_normalise",
    "add_column",
    "flag_anomalies",
    "categorise",
)

# Per-command brief injected into the prompt so one template covers all six.
_COMMAND_BRIEF: dict[str, str] = {
    "explain": (
        "EXPLAIN this selection: what the columns hold and how any formulas work. "
        "Return NO edits — put the explanation in `explanation`."
    ),
    "write_formula": (
        "WRITE or FIX a formula per the user instruction. Return only the formula "
        "edits needed; prefer a single formula that fills the natural target cell."
    ),
    "clean_normalise": (
        "CLEAN & NORMALISE the selected values: trim whitespace, fix obvious casing, "
        "standardise dates and numbers. Return one `value` edit per cell you change."
    ),
    "add_column": (
        "ADD A DERIVED COLUMN immediately to the RIGHT of the selection per the "
        "instruction. Include a header cell and one edit per row; use formulas where "
        "the column is computed."
    ),
    "flag_anomalies": (
        "FLAG ANOMALIES: outliers, sign errors, broken subtotals, duplicates, stale "
        "dates. Return NO edits — list each finding in `warnings` and summarise in "
        "`explanation`."
    ),
    "categorise": (
        "CATEGORISE each row into an accounting category. Add a 'Category' column to "
        "the RIGHT of the selection: one `value` edit per row holding the category, "
        "plus a header edit. Keep categories consistent across similar rows."
    ),
}


class SelectionEdit(BaseModel):
    cell: str                     # absolute A1, within or just right of the selection
    value: str | None = None
    formula: str | None = None    # starts with "="; never set together with value


class SelectionResult(BaseModel):
    edits: list[SelectionEdit] = []
    explanation: str = ""
    warnings: list[str] = []


def _selection_schema() -> dict:
    return {
        "type": "object",
        "properties": {
            "edits": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell": {"type": "string"},
                        "value": {"type": ["string", "null"]},
                        "formula": {"type": ["string", "null"]},
                    },
                    "required": ["cell"],
                },
            },
            "explanation": {"type": "string"},
            "warnings": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["edits", "explanation", "warnings"],
    }


def _selection_spec() -> TaskSpec:
    return TaskSpec(
        name="assistant_selection",
        prompt_version="v1",
        prompt_file="assistant_selection.md",
        output_model=SelectionResult,
        output_schema=_selection_schema(),
        # Deterministic edits, but the grids can be wide — give it more context.
        policy=TaskPolicy(num_ctx=8192),
    )


def _coerce(val: object) -> object:
    """Best-effort string→number so recompute evaluates formulas over real values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    cleaned = s.replace(",", "").replace("₦", "").replace("$", "").strip()
    try:
        f = float(cleaned)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def _render_table(
    selection_a1: str,
    headers: list[str],
    rows: list[list[str]],
    formulas: list[list[str]] | None,
    max_rows: int = 60,
) -> str:
    """Render the selection as a compact table keyed by real A1 column letters and
    row numbers, so the model references cells unambiguously."""
    from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

    start = selection_a1.split(":")[0].strip() or "A1"
    try:
        start_row, start_col = coordinate_to_tuple(start)
    except Exception:
        start_row, start_col = 1, 1

    ncols = max((len(r) for r in rows), default=len(headers))
    ncols = max(ncols, 1)
    col_letters = [get_column_letter(start_col + c) for c in range(ncols)]

    lines = ["      | " + " | ".join(col_letters)]
    if any(h.strip() for h in headers):
        hdr = [headers[c] if c < len(headers) else "" for c in range(ncols)]
        lines.append("(name)| " + " | ".join(hdr))
    for i, row in enumerate(rows[:max_rows]):
        rownum = start_row + i
        cells = []
        for c in range(ncols):
            v = str(row[c]) if c < len(row) else ""
            f = ""
            if formulas and i < len(formulas) and c < len(formulas[i]):
                fx = str(formulas[i][c]).strip()
                if fx:
                    f = f"  [{fx if fx.startswith('=') else '=' + fx}]"
            cells.append(f"{v}{f}")
        lines.append(f"{rownum:>5} | " + " | ".join(cells))
    if len(rows) > max_rows:
        lines.append(f"… ({len(rows) - max_rows} more rows omitted)")
    return "\n".join(lines)


def run_selection_command(
    *,
    command: str,
    selection_a1: str,
    sheet_name: str | None,
    headers: list[str],
    rows: list[list[str]],
    formulas: list[list[str]] | None = None,
    instruction: str | None,
    provider: Provider,
    cache: AiCache | None = None,
    telemetry: AiTelemetry | None = None,
) -> SelectionResult:
    """Ask the model for a structured cell-edit proposal for the selection.

    Degrades gracefully: a provider error or an unparseable answer returns an
    empty proposal with a warning rather than raising (Step 2 safety)."""
    if command not in COMMANDS:
        raise ValueError(f"unknown command {command!r}")

    inputs = {
        "command": command,
        "command_brief": _COMMAND_BRIEF[command],
        "sheet_name": (sheet_name or "Sheet1"),
        "selection_a1": selection_a1,
        "headers": ", ".join(h for h in headers if h.strip()) or "(none)",
        "instruction": (instruction or "").strip() or "(none)",
        "table": _render_table(selection_a1, headers, rows, formulas),
    }
    try:
        result = run_task(
            _selection_spec(), inputs, provider=provider, cache=cache, telemetry=telemetry
        )
    except Exception as exc:  # provider down / network — never 500 the editor
        log.warning("selection command %s failed: %s", command, exc)
        return SelectionResult(
            warnings=[f"The assistant is unavailable right now ({type(exc).__name__})."]
        )
    if result.ok and isinstance(result.model, SelectionResult):
        # An edit must carry exactly one of value/formula; drop empties defensively.
        clean = [e for e in result.model.edits if (e.value is not None) or (e.formula is not None)]
        result.model.edits = clean
        return result.model
    return SelectionResult(
        warnings=["Couldn't produce a confident result — try a smaller selection or rephrase."]
    )


def verify_selection_edits(
    selection_a1: str,
    rows: list[list[str]],
    edits: list[SelectionEdit],
) -> list[str]:
    """Smoke-test proposed FORMULAS: write the selection at its real coordinates
    into a throwaway workbook, apply the edits, recompute through LibreOffice and
    return any error-token warnings. Best-effort — a missing/failing LibreOffice
    is non-fatal (verification degrades to 'unverified', never blocks the edit)."""
    if not any(e.formula for e in edits):
        return []
    try:
        import tempfile
        from pathlib import Path

        import openpyxl
        from openpyxl.utils.cell import (
            coordinate_to_tuple,
            get_column_letter,
        )

        from litchai.validation.recompute import find_workbook_errors, recompute_workbook

        start = selection_a1.split(":")[0].strip() or "A1"
        start_row, start_col = coordinate_to_tuple(start)

        wb = openpyxl.Workbook()
        ws = wb.active
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                ws.cell(row=start_row + r, column=start_col + c, value=_coerce(val))
        for e in edits:
            try:
                er, ec = coordinate_to_tuple(e.cell)
            except Exception:
                continue
            cell = ws.cell(row=er, column=ec)
            if e.formula:
                cell.value = e.formula if e.formula.startswith("=") else "=" + e.formula
            elif e.value is not None:
                cell.value = _coerce(e.value)

        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "selection.xlsx"
            wb.save(path)
            grids = recompute_workbook(path)
        errors = find_workbook_errors(grids)
        return [
            f"{sheet}!{get_column_letter(col + 1)}{row + 1} would error: {token}"
            for (sheet, row, col, token) in errors
        ]
    except Exception as exc:
        log.info("selection verify skipped: %s", exc)
        return [f"Couldn't auto-verify the formula ({type(exc).__name__}); apply with care."]
