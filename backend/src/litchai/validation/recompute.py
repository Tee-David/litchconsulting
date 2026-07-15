"""Headless-LibreOffice recompute gate (PRD step 7 / FR6).

openpyxl never evaluates formulas, so a workbook fresh from a compiler carries
no cached results. LibreOffice computes every formula on load; exporting to
CSV then yields the evaluated grid. Any Excel error token in that grid, or any
mismatch against golden totals, blocks the file before HITL review.
"""
from __future__ import annotations

import csv
import re
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from litchai.compilers._common import split_sheet_ref

ERROR_TOKEN = re.compile(r"#(?:REF|NAME|VALUE|DIV/0|NUM|NULL|N/A)[!?]?|Err:\d+")


class RecomputeError(RuntimeError):
    pass


def recompute_to_grid(xlsx_path: Path, timeout: int = 120) -> list[list[str]]:
    """Convert the first sheet to CSV via headless LibreOffice, forcing formula
    evaluation, and return the grid as rows of strings."""
    with tempfile.TemporaryDirectory() as tmp:
        profile = Path(tmp) / "profile"
        cmd = [
            "soffice",
            "--headless",
            "--norestore",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to",
            "csv",
            "--outdir",
            tmp,
            str(xlsx_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        out_csv = Path(tmp) / (xlsx_path.stem + ".csv")
        if proc.returncode != 0 or not out_csv.exists():
            raise RecomputeError(
                f"LibreOffice conversion failed (rc={proc.returncode}): {proc.stderr.strip()}"
            )
        with out_csv.open(newline="", encoding="utf-8") as fh:
            return list(csv.reader(fh))


def recompute_workbook(xlsx_path: Path, timeout: int = 180) -> dict[str, list[list[str]]]:
    """Round-trip the workbook through headless LibreOffice (xlsx -> xlsx),
    forcing every formula on every sheet to recompute, and return each sheet as
    a grid of strings keyed by sheet name.

    CSV conversion (recompute_to_grid) only exports the first sheet; the xlsx
    round-trip keeps sheet names intact and writes a cached value for every
    formula cell. Compiler-fresh workbooks carry no cached values, so every
    number in the round-tripped file was necessarily computed by LibreOffice.
    """
    with tempfile.TemporaryDirectory() as tmp:
        profile = Path(tmp) / "profile"
        cmd = [
            "soffice",
            "--headless",
            "--norestore",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            tmp,
            str(xlsx_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        out_xlsx = Path(tmp) / (xlsx_path.stem + ".xlsx")
        if proc.returncode != 0 or not out_xlsx.exists():
            raise RecomputeError(
                f"LibreOffice conversion failed (rc={proc.returncode}): {proc.stderr.strip()}"
            )
        wb = load_workbook(out_xlsx, data_only=True, read_only=True)
        try:
            return {
                ws.title: [
                    ["" if v is None else str(v) for v in row]
                    for row in ws.iter_rows(values_only=True)
                ]
                for ws in wb.worksheets
            }
        finally:
            wb.close()


def value_at_ref(grids: dict[str, list[list[str]]], qualified: str) -> float:
    """Read a numeric value at a sheet-qualified ref ("SOFP!D26",
    "'Bank Recon'!F55") from a recompute_workbook result."""
    sheet, ref = split_sheet_ref(qualified)
    if sheet is None:
        raise RecomputeError(f"ref {qualified!r} carries no sheet qualifier")
    if sheet not in grids:
        raise RecomputeError(f"sheet {sheet!r} not in workbook (has {sorted(grids)})")
    return value_at(grids[sheet], ref)


def find_workbook_errors(grids: dict[str, list[list[str]]]) -> list[tuple[str, int, int, str]]:
    """(sheet, row, col, content) for every error-token cell across all sheets."""
    return [
        (name, r, c, content)
        for name, grid in grids.items()
        for r, c, content in find_error_cells(grid)
    ]


def find_error_cells(grid: list[list[str]]) -> list[tuple[int, int, str]]:
    """(row, col, content) for every cell containing an Excel error token; 1-indexed."""
    hits = []
    for r, row in enumerate(grid, start=1):
        for c, cell in enumerate(row, start=1):
            if ERROR_TOKEN.search(cell):
                hits.append((r, c, cell))
    return hits


def parse_amount(cell: str) -> float:
    """Parse a CSV cell that may carry currency formatting (₦, thousands
    separators, parenthesised negatives)."""
    cleaned = cell.replace("₦", "").replace(",", "").replace(" ", "").strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    return float(cleaned)


def cell_ref_to_index(ref: str) -> tuple[int, int]:
    """'C12' -> (12, 3); 1-indexed row and column."""
    m = re.fullmatch(r"([A-Z]+)(\d+)", ref)
    if not m:
        raise ValueError(f"bad cell ref: {ref!r}")
    col = 0
    for ch in m.group(1):
        col = col * 26 + ord(ch) - 64
    return int(m.group(2)), col


def value_at(grid: list[list[str]], ref: str) -> float:
    row, col = cell_ref_to_index(ref)
    try:
        return parse_amount(grid[row - 1][col - 1])
    except (IndexError, ValueError) as exc:
        raise RecomputeError(f"cannot read numeric value at {ref}: {exc}") from exc
