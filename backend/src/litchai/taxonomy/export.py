"""Excel round-trip for operator sign-off: the firm reviews its chart of
categories as a spreadsheet, not JSON. Import covers category edits;
migrations remain JSON-managed (they encode history, not current state)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from litchai.compilers._common import BOLD, TITLE
from litchai.compilers.annual_report._rows import write_text
from litchai.taxonomy.loader import _validate
from litchai.taxonomy.model import Category, Taxonomy

_HEADERS = ("code", "label", "parent", "postable", "nature", "source", "keywords", "template_labels")
_HEADER_ROW = 4


def export_taxonomy_xlsx(taxonomy: Taxonomy, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Taxonomy"
    ws.sheet_view.showGridLines = False
    widths = {"A": 32, "B": 44, "C": 22, "D": 10, "E": 14, "F": 12, "G": 50, "H": 50}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    write_text(ws, "A1", f"LitchAI category taxonomy — version {taxonomy.version}", font=TITLE)
    write_text(ws, "A2", "Derived from: " + "; ".join(taxonomy.derived_from))
    for col_index, header in enumerate(_HEADERS, start=1):
        write_text(ws, f"{chr(64 + col_index)}{_HEADER_ROW}", header, font=BOLD)

    for row_index, category in enumerate(taxonomy.categories, start=_HEADER_ROW + 1):
        values = (
            category.code,
            category.label,
            category.parent or "",
            "TRUE" if category.postable else "FALSE",
            category.nature or "",
            category.source,
            ", ".join(category.keywords),
            " | ".join(category.template_labels),
        )
        for col_index, value in enumerate(values, start=1):
            write_text(ws, f"{chr(64 + col_index)}{row_index}", value)

    wb.save(path)


def import_taxonomy_xlsx(path: Path) -> Taxonomy:
    ws = load_workbook(path, data_only=True)["Taxonomy"]
    title = str(ws["A1"].value or "")
    version = title.rsplit("version ", 1)[-1].strip()
    derived = str(ws["A2"].value or "").removeprefix("Derived from: ")

    categories = []
    for row in ws.iter_rows(min_row=_HEADER_ROW + 1, max_col=len(_HEADERS)):
        raw = {h: (c.value if c.value is not None else "") for h, c in zip(_HEADERS, row)}
        if not str(raw["code"]).strip():
            continue
        categories.append(
            Category(
                code=str(raw["code"]).strip(),
                label=str(raw["label"]),
                parent=str(raw["parent"]).strip() or None,
                postable=str(raw["postable"]).strip().upper() == "TRUE",
                nature=str(raw["nature"]).strip() or None,
                source=str(raw["source"]).strip() or "transactions",
                keywords=[k.strip() for k in str(raw["keywords"]).split(",") if k.strip()],
                template_labels=[
                    t.strip() for t in str(raw["template_labels"]).split("|") if t.strip()
                ],
            )
        )

    taxonomy = Taxonomy(
        version=version,
        derived_from=[d.strip() for d in derived.split(";") if d.strip()],
        categories=categories,
    )
    _validate(taxonomy)
    return taxonomy
