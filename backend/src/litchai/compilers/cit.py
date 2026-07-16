"""Company Income Tax Computation & Capital Allowance Register compiler (v1.1).

Deterministic (PRD §3): the NTA 2025 rates and small-company thresholds are
written from the shared tax config into labelled input cells, and *every*
computed figure — capital allowances, assessable profit, the small-company
determination, the CIT rate choice and the Development Levy — is an Excel
formula. The small-company test itself is an ``IF(AND(...))`` over the threshold
cells, so the 0%-vs-30% decision is never hardcoded.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    SUBTLE,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.cit import AdjLine, CitContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "cit-1.0.0"

LABEL_COL = "B"
C = "C"  # amount / value
D = "D"  # capital-allowance rate
E = "E"  # capital-allowance amount


def compile_cit(contract: CitContract) -> CompiledTemplate:
    cfg = load_tax_config()
    cit = cfg["cit"]

    wb = Workbook()
    ws = wb.active
    ws.title = "CIT Computation"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 46
    for col in (C, D, E):
        ws.column_dimensions[col].width = 18

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "Company Income Tax Computation"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.period_label

    key: dict[str, str] = {}
    row = 5

    def header(text: str) -> None:
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = text
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1

    def value_row(label: str, value, fmt: str = NAIRA_FMT, col: str = C) -> str:
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = label
        ref = f"{col}{row}"
        cell = ws[ref]
        cell.value = value
        cell.number_format = fmt
        row += 1
        return ref

    def section(title: str, items: list[AdjLine]) -> str:
        nonlocal row
        header(title)
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = f"    {item.label}"
            cell = ws[f"{C}{row}"]
            cell.value = item.amount
            cell.number_format = NAIRA_FMT
            row += 1
        ref = f"{C}{row}"
        ws[f"{LABEL_COL}{row}"] = "    Subtotal"
        sub = ws[ref]
        sub.value = f"=SUM({C}{first}:{C}{row - 1})" if items else 0
        sub.number_format = NAIRA_FMT
        sub.border = TOP_BORDER
        row += 1
        return ref

    def result_row(label: str, formula: str, col: str = C) -> str:
        nonlocal row
        ref = f"{col}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = formula
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = DOUBLE_TOP
        row += 2
        return ref

    # Tax parameters (from the shared config — inputs, not computed).
    header("Tax parameters (NTA 2025)")
    std_rate = value_row("Standard CIT rate", cit["standardRatePct"] / 100.0, "0%")
    small_rate = value_row("Small-company CIT rate", cit["smallCompany"]["ratePct"] / 100.0, "0%")
    levy_rate = value_row("Development Levy rate", cit["developmentLevy"]["ratePct"] / 100.0, "0%")
    max_turnover = value_row("Small-company max turnover", cit["smallCompany"]["maxTurnover"])
    max_assets = value_row("Small-company max fixed assets", cit["smallCompany"]["maxFixedAssets"])
    row += 1

    # Small-company test — the determination is a formula over the inputs.
    header("Small-company test")
    turnover_ref = value_row("Gross turnover", contract.turnover)
    assets_ref = value_row("Total fixed assets", contract.total_fixed_assets)
    prof_ref = value_row(
        "Professional services (1 = yes)", 1 if contract.professional_services else 0, "0"
    )
    is_small = result_row(
        "Small company? (1 = yes)",
        f"=IF(AND({turnover_ref}<={max_turnover},{assets_ref}<={max_assets},{prof_ref}=0),1,0)",
    )
    key["is_small"] = is_small

    # Capital Allowance Register.
    header("Capital Allowance Register")
    ws[f"{C}{row}"] = "Cost"
    ws[f"{D}{row}"] = "Rate"
    ws[f"{E}{row}"] = "Allowance"
    for col in (C, D, E):
        ws[f"{col}{row}"].font = BOLD
    row += 1
    ca_first = row
    for asset in contract.capital_allowance_assets:
        ws[f"{LABEL_COL}{row}"] = f"    {asset.description}"
        cost = ws[f"{C}{row}"]
        cost.value = asset.cost
        cost.number_format = NAIRA_FMT
        rate = ws[f"{D}{row}"]
        rate.value = asset.allowance_rate_pct / 100.0
        rate.number_format = "0%"
        allowance = ws[f"{E}{row}"]
        allowance.value = f"={C}{row}*{D}{row}"
        allowance.number_format = NAIRA_FMT
        row += 1
    ws[f"{LABEL_COL}{row}"] = "Total capital allowances"
    ws[f"{LABEL_COL}{row}"].font = BOLD
    total_ca = ws[f"{E}{row}"]
    total_ca.value = f"=SUM({E}{ca_first}:{E}{row - 1})" if contract.capital_allowance_assets else 0
    total_ca.font = BOLD
    total_ca.number_format = NAIRA_FMT
    total_ca.border = TOP_BORDER
    total_ca_ref = f"{E}{row}"
    key["total_capital_allowances"] = total_ca_ref
    row += 2

    # CIT computation.
    header("CIT computation")
    net_profit = value_row("Net profit per accounts", contract.net_profit_per_accounts)
    addbacks = section("Add: disallowable expenses", contract.disallowable_addbacks)
    ca_line = value_row("Less: capital allowances", f"={total_ca_ref}")
    deductions = section("Less: other allowable deductions", contract.other_deductions)
    assessable = result_row(
        "Assessable profit",
        f"={net_profit}+{addbacks}-{ca_line}-{deductions}",
    )
    key["assessable_profit"] = assessable
    key["cit_payable"] = result_row(
        "CIT payable",
        f"={assessable}*IF({is_small}=1,{small_rate},{std_rate})",
    )
    key["development_levy"] = result_row(
        "Development Levy",
        f"=IF({is_small}=1,0,{assessable}*{levy_rate})",
    )
    key["total_tax"] = result_row(
        "Total tax payable",
        f"={key['cit_payable']}+{key['development_levy']}",
    )

    ws[f"{LABEL_COL}{row}"] = (
        "Small companies (≤₦100m turnover ∧ ≤₦250m assets, non-professional) pay 0% CIT and no levy."
    )
    ws[f"{LABEL_COL}{row}"].font = SUBTLE
    row += 2

    write_footer(ws, f"{LABEL_COL}{row}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
