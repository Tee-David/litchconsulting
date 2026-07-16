"""Withholding Tax Schedule & Credit Note Reconciliation compiler (v1.1).

Deterministic (PRD §3): gross amounts are inputs and each WHT rate is written
from the shared tax config into a per-line rate cell; the WHT deducted is
``gross * rate`` as a formula, exempt lines force a 0% rate cell. The
reconciliation control — total WHT deducted vs total credit notes issued —
compiles to a difference that should be 0 (a non-zero value is the sheet
correctly reporting a missing credit note).
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    SUBTLE,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.wht import WhtScheduleContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "wht-1.0.0"

VENDOR_COL = "B"
CATEGORY_COL = "C"
GROSS_COL = "D"
RATE_COL = "E"
WHT_COL = "F"


def _rate_for(cfg: dict, category: str, payee_type: str) -> float:
    rate = cfg["wht"]["rates"].get(category)
    if rate is None:
        return 0.0
    return rate[f"{payee_type}Pct"] / 100.0


def compile_wht_schedule(contract: WhtScheduleContract) -> CompiledTemplate:
    cfg = load_tax_config()
    exemption = cfg["wht"]["smallSupplierExemption"]["maxMonthlyPayment"]

    wb = Workbook()
    ws = wb.active
    ws.title = "WHT Schedule"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[VENDOR_COL].width = 34
    ws.column_dimensions[CATEGORY_COL].width = 24
    for col in (GROSS_COL, RATE_COL, WHT_COL):
        ws.column_dimensions[col].width = 16

    ws[f"{VENDOR_COL}1"] = contract.client_name
    ws[f"{VENDOR_COL}1"].font = TITLE
    ws[f"{VENDOR_COL}2"] = "WHT Schedule & Credit Note Reconciliation"
    ws[f"{VENDOR_COL}2"].font = BOLD
    ws[f"{VENDOR_COL}3"] = contract.period_label

    for col, head in ((VENDOR_COL, "Vendor"), (CATEGORY_COL, "Category"),
                      (GROSS_COL, "Gross"), (RATE_COL, "Rate"), (WHT_COL, "WHT")):
        ws[f"{col}5"] = head
        ws[f"{col}5"].font = BOLD

    key: dict[str, str] = {}
    row = 6
    first = row
    for line in contract.deductions:
        rate = 0.0 if line.exempt else _rate_for(cfg, line.category, line.payee_type)
        label = cfg["wht"]["rates"].get(line.category, {}).get("label", line.category)
        ws[f"{VENDOR_COL}{row}"] = line.vendor
        ws[f"{CATEGORY_COL}{row}"] = f"{label}{' (exempt)' if line.exempt else ''}"
        gross = ws[f"{GROSS_COL}{row}"]
        gross.value = line.gross_amount
        gross.number_format = NAIRA_FMT
        rate_cell = ws[f"{RATE_COL}{row}"]
        rate_cell.value = rate
        rate_cell.number_format = "0%"
        wht = ws[f"{WHT_COL}{row}"]
        wht.value = f"={GROSS_COL}{row}*{RATE_COL}{row}"
        wht.number_format = NAIRA_FMT
        row += 1

    total_row = row
    ws[f"{VENDOR_COL}{total_row}"] = "Total WHT deducted"
    ws[f"{VENDOR_COL}{total_row}"].font = BOLD
    deducted = ws[f"{WHT_COL}{total_row}"]
    deducted.value = f"=SUM({WHT_COL}{first}:{WHT_COL}{row - 1})" if contract.deductions else 0
    deducted.font = BOLD
    deducted.number_format = NAIRA_FMT
    deducted.border = TOP_BORDER
    key["total_wht_deducted"] = f"{WHT_COL}{total_row}"
    row += 2

    # Credit notes issued for those deductions.
    ws[f"{VENDOR_COL}{row}"] = "Credit notes issued"
    ws[f"{VENDOR_COL}{row}"].font = BOLD
    row += 1
    cn_first = row
    for note in contract.credit_notes:
        ws[f"{VENDOR_COL}{row}"] = f"    {note.vendor}".rstrip()
        ws[f"{CATEGORY_COL}{row}"] = note.reference
        amt = ws[f"{WHT_COL}{row}"]
        amt.value = note.amount
        amt.number_format = NAIRA_FMT
        row += 1
    ws[f"{VENDOR_COL}{row}"] = "Total credit notes issued"
    ws[f"{VENDOR_COL}{row}"].font = BOLD
    credits = ws[f"{WHT_COL}{row}"]
    credits.value = f"=SUM({WHT_COL}{cn_first}:{WHT_COL}{row - 1})" if contract.credit_notes else 0
    credits.font = BOLD
    credits.number_format = NAIRA_FMT
    credits.border = TOP_BORDER
    key["total_credit_notes"] = f"{WHT_COL}{row}"
    row += 2

    ws[f"{VENDOR_COL}{row}"] = "Difference (should be 0)"
    ws[f"{VENDOR_COL}{row}"].font = BOLD
    diff = ws[f"{WHT_COL}{row}"]
    diff.value = f"={key['total_wht_deducted']}-{key['total_credit_notes']}"
    diff.font = BOLD
    diff.number_format = NAIRA_FMT
    diff.border = DOUBLE_TOP
    key["difference"] = f"{WHT_COL}{row}"
    row += 2

    ws[f"{VENDOR_COL}{row}"] = (
        f"Small-supplier exemption: payments ≤ ₦{exemption:,.0f}/month may be exempt (config)."
    )
    ws[f"{VENDOR_COL}{row}"].font = SUBTLE
    row += 2

    write_footer(ws, f"{VENDOR_COL}{row}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
