"""Cash Flow Statement compiler — Direct method.

Deterministic (PRD §3): each receipt/payment and the opening cash balance are
the raw inputs; every subtotal, each activity's net cash, the net change, and
closing cash are Excel formulas.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.cashflow import CashflowContract, CashLine
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "cashflow-1.0.0"

LABEL_COL = "B"
AMOUNT_COL = "C"


def compile_cashflow(contract: CashflowContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 46
    ws.column_dimensions[AMOUNT_COL].width = 18

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "Statement of Cash Flows (Direct method)"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.period_label

    key: dict[str, str] = {}
    row = 5

    def sub_block(title: str, items: list[CashLine]) -> str:
        """List items under a sub-heading and return the subtotal cell ref."""
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        row += 1
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = f"    {item.label}"
            cell = ws[f"{AMOUNT_COL}{row}"]
            cell.value = item.amount
            cell.number_format = NAIRA_FMT
            row += 1
        ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = "    Subtotal"
        sub = ws[ref]
        sub.value = f"=SUM({AMOUNT_COL}{first}:{AMOUNT_COL}{row - 1})" if items else 0
        sub.number_format = NAIRA_FMT
        sub.border = TOP_BORDER
        row += 1
        return ref

    def activity(title: str, receipts: list[CashLine], payments: list[CashLine]) -> str:
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1
        receipts_ref = sub_block("Receipts", receipts)
        payments_ref = sub_block("Payments", payments)
        net_ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = f"Net cash from {title.split('from ')[-1]}"
        ws[f"{LABEL_COL}{row}"].font = BOLD
        net = ws[net_ref]
        net.value = f"={receipts_ref}-{payments_ref}"
        net.font = BOLD
        net.number_format = NAIRA_FMT
        net.border = TOP_BORDER
        row += 2
        return net_ref

    key["net_operating"] = activity(
        "Cash flows from operating activities",
        contract.operating_receipts,
        contract.operating_payments,
    )
    key["net_investing"] = activity(
        "Cash flows from investing activities",
        contract.investing_receipts,
        contract.investing_payments,
    )
    key["net_financing"] = activity(
        "Cash flows from financing activities",
        contract.financing_receipts,
        contract.financing_payments,
    )

    # Net change, opening, closing.
    net_change_ref = f"{AMOUNT_COL}{row}"
    ws[f"{LABEL_COL}{row}"] = "Net increase / (decrease) in cash"
    ws[f"{LABEL_COL}{row}"].font = BOLD
    net_change = ws[net_change_ref]
    net_change.value = f"={key['net_operating']}+{key['net_investing']}+{key['net_financing']}"
    net_change.font = BOLD
    net_change.number_format = NAIRA_FMT
    net_change.border = TOP_BORDER
    key["net_change"] = net_change_ref
    row += 1

    opening_ref = f"{AMOUNT_COL}{row}"
    ws[f"{LABEL_COL}{row}"] = "Cash at beginning of period"
    ws[opening_ref] = contract.opening_cash
    ws[opening_ref].number_format = NAIRA_FMT
    row += 1

    closing_ref = f"{AMOUNT_COL}{row}"
    ws[f"{LABEL_COL}{row}"] = "Cash at end of period"
    ws[f"{LABEL_COL}{row}"].font = BOLD
    closing = ws[closing_ref]
    closing.value = f"={opening_ref}+{net_change_ref}"
    closing.font = BOLD
    closing.number_format = NAIRA_FMT
    closing.border = DOUBLE_TOP
    key["closing_cash"] = closing_ref
    row += 2

    write_footer(ws, f"{LABEL_COL}{row}", COMPILER_VERSION, cfg["version"])

    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
