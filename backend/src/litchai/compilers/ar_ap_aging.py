"""Accounts Receivable / Payable Aging compiler (v1.1).

Deterministic (PRD §3): amounts and dates are inputs; each item's age
(``as_at − due_date``) and its placement into the Not-due / 1–30 / 31–60 / 61–90 /
90+ buckets are Excel formulas, and the bucket totals reconcile to the grand
total (a built-in ``check`` cell that is 0 when every naira is bucketed once).
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.ar_ap_aging import AgingContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "araping-1.0.0"

PARTY, REF, AMT, DUE, DAYS = "B", "C", "D", "E", "F"
BUCKETS = ["G", "H", "I", "J", "K"]  # not-due, 1-30, 31-60, 61-90, 90+
BUCKET_LABELS = ["Not due", "1–30", "31–60", "61–90", "90+"]
DATE_FMT = "yyyy-mm-dd"


def compile_ar_ap_aging(contract: AgingContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Aging"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[PARTY].width = 26
    ws.column_dimensions[REF].width = 16
    for col in (AMT, DUE, DAYS, *BUCKETS):
        ws.column_dimensions[col].width = 13

    ws[f"{PARTY}1"] = contract.client_name
    ws[f"{PARTY}1"].font = TITLE
    noun = "Debtors" if contract.kind == "receivables" else "Creditors"
    ws[f"{PARTY}2"] = f"{noun} Aging ({contract.kind})"
    ws[f"{PARTY}2"].font = BOLD
    ws[f"{PARTY}3"] = contract.period_label

    ws[f"{PARTY}4"] = "As at"
    ws[f"{PARTY}4"].font = BOLD
    as_at_ref = f"{REF}4"
    ws[as_at_ref] = contract.as_at
    ws[as_at_ref].number_format = DATE_FMT

    headers = ["Party", "Ref", "Amount", "Due", "Days", *BUCKET_LABELS]
    for col, head in zip([PARTY, REF, AMT, DUE, DAYS, *BUCKETS], headers):
        ws[f"{col}6"] = head
        ws[f"{col}6"].font = BOLD

    key: dict[str, str] = {}
    row = 7
    first = row
    for item in contract.items:
        ws[f"{PARTY}{row}"] = item.party
        ws[f"{REF}{row}"] = item.reference
        amt = ws[f"{AMT}{row}"]
        amt.value = item.amount
        amt.number_format = NAIRA_FMT
        due = ws[f"{DUE}{row}"]
        due.value = item.due_date
        due.number_format = DATE_FMT
        days = ws[f"{DAYS}{row}"]
        days.value = f"={as_at_ref}-{DUE}{row}"
        days.number_format = "0"
        d = f"{DAYS}{row}"
        a = f"{AMT}{row}"
        formulas = [
            f"=IF({d}<=0,{a},0)",
            f"=IF(AND({d}>0,{d}<=30),{a},0)",
            f"=IF(AND({d}>30,{d}<=60),{a},0)",
            f"=IF(AND({d}>60,{d}<=90),{a},0)",
            f"=IF({d}>90,{a},0)",
        ]
        for col, formula in zip(BUCKETS, formulas):
            cell = ws[f"{col}{row}"]
            cell.value = formula
            cell.number_format = NAIRA_FMT
        row += 1

    last = row - 1
    ws[f"{PARTY}{row}"] = "Total"
    ws[f"{PARTY}{row}"].font = BOLD
    grand = ws[f"{AMT}{row}"]
    grand.value = f"=SUM({AMT}{first}:{AMT}{last})" if contract.items else 0
    grand.font = BOLD
    grand.number_format = NAIRA_FMT
    grand.border = TOP_BORDER
    key["grand_total"] = f"{AMT}{row}"
    for i, col in enumerate(BUCKETS):
        cell = ws[f"{col}{row}"]
        cell.value = f"=SUM({col}{first}:{col}{last})" if contract.items else 0
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = TOP_BORDER
        key[f"total_{['not_due', 'b1_30', 'b31_60', 'b61_90', 'b90_plus'][i]}"] = f"{col}{row}"
    total_row = row
    row += 2

    ws[f"{PARTY}{row}"] = "Bucket check (should be 0)"
    ws[f"{PARTY}{row}"].font = BOLD
    check = ws[f"{AMT}{row}"]
    bucket_sum = "+".join(f"{col}{total_row}" for col in BUCKETS)
    check.value = f"={AMT}{total_row}-({bucket_sum})"
    check.font = BOLD
    check.number_format = NAIRA_FMT
    check.border = DOUBLE_TOP
    key["bucket_check"] = f"{AMT}{row}"
    row += 2

    write_footer(ws, f"{PARTY}{row}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
