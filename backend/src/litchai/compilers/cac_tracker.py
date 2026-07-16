"""CAC Annual Return & Statutory Compliance Tracker compiler (v1.1).

Deterministic (PRD §3): the requirement rows (due date, completed flag, fee) are
inputs; the completion %, outstanding/overdue counts and outstanding fees are
formulas (``SUM`` / ``SUMPRODUCT`` over the flag and date columns), and a per-row
status is an ``IF`` over the completed flag and the as-at date.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.cac_tracker import CacTrackerContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "cac-1.0.0"

REQ, AUTH, DUE, DONE, STATUS, FEE = "B", "C", "D", "E", "F", "G"
DATE_FMT = "yyyy-mm-dd"


def compile_cac_tracker(contract: CacTrackerContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[REQ].width = 38
    ws.column_dimensions[AUTH].width = 12
    for col in (DUE, DONE, STATUS, FEE):
        ws.column_dimensions[col].width = 13

    ws[f"{REQ}1"] = contract.client_name
    ws[f"{REQ}1"].font = TITLE
    ws[f"{REQ}2"] = "CAC Annual Return & Statutory Compliance Tracker"
    ws[f"{REQ}2"].font = BOLD
    ws[f"{REQ}3"] = contract.period_label

    ws[f"{REQ}4"] = "As at"
    ws[f"{REQ}4"].font = BOLD
    as_at_ref = f"{AUTH}4"
    ws[as_at_ref] = contract.as_at
    ws[as_at_ref].number_format = DATE_FMT
    meta = " · ".join(p for p in (
        f"RC {contract.rc_number}" if contract.rc_number else "",
        f"TIN {contract.tin}" if contract.tin else "",
    ) if p)
    if meta:
        ws[f"{DUE}4"] = meta

    for col, head in ((REQ, "Requirement"), (AUTH, "Authority"), (DUE, "Due"),
                      (DONE, "Done (1/0)"), (STATUS, "Status"), (FEE, "Fee")):
        ws[f"{col}6"] = head
        ws[f"{col}6"].font = BOLD

    row = 7
    first = row
    for item in contract.items:
        ws[f"{REQ}{row}"] = item.requirement
        ws[f"{AUTH}{row}"] = item.authority
        due = ws[f"{DUE}{row}"]
        due.value = item.due_date
        due.number_format = DATE_FMT
        done = ws[f"{DONE}{row}"]
        done.value = 1 if item.completed else 0
        done.number_format = "0"
        status = ws[f"{STATUS}{row}"]
        status.value = (
            f'=IF({DONE}{row}=1,"Filed",IF({as_at_ref}>{DUE}{row},"Overdue","Pending"))'
        )
        fee = ws[f"{FEE}{row}"]
        fee.value = item.fee
        fee.number_format = NAIRA_FMT
        row += 1

    last = row - 1
    dn = f"{DONE}{first}:{DONE}{last}"
    du = f"{DUE}{first}:{DUE}{last}"
    fr = f"{FEE}{first}:{FEE}{last}"
    n = len(contract.items)

    key: dict[str, str] = {}
    row += 1

    def summary(label: str, formula, fmt: str = "0") -> str:
        nonlocal row
        ws[f"{REQ}{row}"] = label
        ws[f"{REQ}{row}"].font = BOLD
        ref = f"{FEE}{row}"
        cell = ws[ref]
        cell.value = formula
        cell.number_format = fmt
        cell.font = BOLD
        row += 1
        return ref

    key["total_items"] = summary("Total requirements", f"=COUNT({dn})" if n else 0)
    key["completed_count"] = summary("Completed", f"=SUM({dn})" if n else 0)
    key["outstanding_count"] = summary(
        "Outstanding", f"={key['total_items']}-{key['completed_count']}"
    )
    key["overdue_count"] = summary(
        "Overdue", f"=SUMPRODUCT(({dn}=0)*({as_at_ref}>{du}))" if n else 0
    )
    key["completion_pct"] = summary(
        "Completion %",
        f"={key['completed_count']}/{key['total_items']}*100" if n else 0,
        fmt="0.0",
    )
    key["total_fees"] = summary("Total fees", f"=SUM({fr})" if n else 0, fmt=NAIRA_FMT)
    key["outstanding_fees"] = summary(
        "Outstanding fees", f"=SUMPRODUCT(({dn}=0)*({fr}))" if n else 0, fmt=NAIRA_FMT
    )
    ws[f"{REQ}{last + 2}"].border = TOP_BORDER  # rule above the summary block

    row += 1
    write_footer(ws, f"{REQ}{row}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
