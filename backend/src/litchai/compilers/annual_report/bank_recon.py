"""Monthly bank-reconciliation sheet (template section structure 1–7 + sign-off).

The adjusted-books cell is exported through a dedicated feed cell (column F,
like the template's F55) that the SOFP builder links as cash and equivalents;
the difference check must recompute to zero.
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet

from litchai.compilers._common import BOLD, THOUSANDS_FMT, TITLE, sheet_ref
from litchai.compilers.annual_report._rows import CUR_COL, LABEL_COL, SheetWriter, write_text
from litchai.contracts.annual_report import BankReconSection, DatedLine

SHEET = "Bank Recon"
FEED_COL = "F"


@dataclass(frozen=True)
class BankReconRefs:
    feed: str
    adjusted_bank: str
    adjusted_books: str
    difference_check: str


def _dated_section(
    w: SheetWriter, ws: Worksheet, title: str, lines: list[DatedLine], total_label: str
) -> int:
    w.section(title)
    write_text(ws, f"{LABEL_COL}{w.row}", "Date", font=BOLD)
    write_text(ws, f"C{w.row}", "Description", font=BOLD)
    write_text(ws, f"{CUR_COL}{w.row}", "Amount", font=BOLD)
    w.row += 1
    first = w.row
    for line in lines:
        write_text(ws, f"{LABEL_COL}{w.row}", line.date)
        write_text(ws, f"C{w.row}", line.description)
        cell = ws[f"{CUR_COL}{w.row}"]
        cell.value = line.amount
        cell.number_format = THOUSANDS_FMT
        w.row += 1
    cur, _ = w.sum_rows(first, w.row - 1)
    total_row = w.formula_row(total_label, cur, bold=True, border="top")
    w.blank()
    return total_row


def _amount_row(w: SheetWriter, ws: Worksheet, label: str, amount: float) -> int:
    write_text(ws, f"{LABEL_COL}{w.row}", label)
    cell = ws[f"{CUR_COL}{w.row}"]
    cell.value = amount
    cell.number_format = THOUSANDS_FMT
    w.row += 1
    return w.row - 1


def build_bank_recon(
    ws: Worksheet, contract, key_cells: dict[str, str]
) -> BankReconRefs:
    recon: BankReconSection = contract.bank_recon
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions[LABEL_COL].width = 16
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions[CUR_COL].width = 16
    ws.column_dimensions[FEED_COL].width = 16

    w = SheetWriter(ws)
    write_text(ws, "A1", contract.client_name, font=TITLE)
    write_text(ws, "A2", "Monthly Bank Reconciliation", font=BOLD)
    write_text(ws, "A3", recon.account_label)
    w.row = 5

    w.section("1. Balance per bank statement")
    stmt_row = _amount_row(
        w, ws, "Closing balance per bank statement, as at month-end", recon.statement_balance
    )
    w.blank()

    deposits_total = _dated_section(
        w, ws,
        "2. Add: Deposits in transit (recorded in books, not yet on bank statement)",
        recon.deposits_in_transit, "Total deposits in transit",
    )
    cheques_total = _dated_section(
        w, ws,
        "3. Less: Outstanding/unpresented cheques and payments",
        recon.outstanding_cheques, "Total outstanding cheques/payments",
    )
    adjusted_bank = w.formula_row(
        "Adjusted balance per bank (should equal adjusted balance per books below)",
        f"={CUR_COL}{stmt_row}+{CUR_COL}{deposits_total}-{CUR_COL}{cheques_total}",
        bold=True, border="top",
    )
    w.key("bank_recon:adjusted_bank", adjusted_bank, prior=False)
    w.blank()

    w.section("4. Balance per cash book / general ledger")
    book_row = _amount_row(
        w, ws, "Closing balance per cash book, as at month-end", recon.book_balance
    )
    w.blank()

    credits_total = _dated_section(
        w, ws,
        "5. Add: Unrecorded bank credits (e.g. direct deposits, interest earned, standing orders)",
        recon.unrecorded_credits, "Total unrecorded credits",
    )
    debits_total = _dated_section(
        w, ws,
        "6. Less: Unrecorded bank charges/debits (e.g. bank charges, COT, failed cheques, stamp duty)",
        recon.unrecorded_debits, "Total unrecorded charges/debits",
    )
    adjusted_books = w.formula_row(
        "Adjusted balance per cash book",
        f"={CUR_COL}{book_row}+{CUR_COL}{credits_total}-{CUR_COL}{debits_total}",
        bold=True, border="top",
    )
    w.key("bank_recon:adjusted_books", adjusted_books, prior=False)
    w.blank()

    w.section("7. Reconciliation check")
    check_row = w.formula_row(
        "Difference (adjusted bank − adjusted books; must equal zero before sign-off)",
        f"={CUR_COL}{adjusted_bank}-{CUR_COL}{adjusted_books}",
    )
    w.key("bank_recon:difference_check", check_row, prior=False)
    feed_row = w.row
    write_text(
        ws, f"{LABEL_COL}{feed_row}",
        "Adjusted cash balance (feeds Statement of Financial Position)",
    )
    feed_cell = ws[f"{FEED_COL}{feed_row}"]
    feed_cell.value = f"={CUR_COL}{adjusted_books}"
    feed_cell.number_format = THOUSANDS_FMT
    feed_cell.font = BOLD
    w.key_cells["bank_recon:feed"] = f"{FEED_COL}{feed_row}"
    w.row += 2

    w.section("Sign-off")
    for label, value in (
        ("Prepared by:", recon.prepared_by),
        ("Date prepared:", recon.prepared_date),
        ("Reviewed/approved by:", recon.reviewed_by),
        ("Date reviewed:", recon.reviewed_date),
    ):
        write_text(ws, f"{LABEL_COL}{w.row}", label)
        write_text(ws, f"{CUR_COL}{w.row}", value)
        w.row += 1

    for name, local_ref in w.key_cells.items():
        key_cells[name] = sheet_ref(SHEET, local_ref)

    return BankReconRefs(
        feed=key_cells["bank_recon:feed"],
        adjusted_bank=key_cells["bank_recon:adjusted_bank"],
        adjusted_books=key_cells["bank_recon:adjusted_books"],
        difference_check=key_cells["bank_recon:difference_check"],
    )
