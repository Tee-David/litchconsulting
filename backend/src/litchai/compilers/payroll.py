"""PAYE & Statutory Payroll Run compiler — the first compiler to consume the
shared Nigerian tax config.

Deterministic (PRD §3): only each employee's gross is a raw input. Pension, NHF,
chargeable income, **PAYE (progressive, generated from the config bands)**, net
pay, and the employer contributions are all Excel formulas. Rates come from
`nigeria-tax-config.json`, never hardcoded (PRD §12).
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.payroll import PayrollContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "payroll-1.0.0"

NAME_COL = "B"
GROSS_COL = "C"
PENSION_COL = "D"
NHF_COL = "E"
CHARGEABLE_COL = "F"
PAYE_COL = "G"
NET_COL = "H"
EMP_PENSION_COL = "I"
NSITF_COL = "J"


def paye_formula(bands: list[dict], ref: str) -> str:
    """Progressive PAYE as an Excel formula over the chargeable-income cell.

    tax = Σ rateᵢ·MIN(MAX(C−Lᵢ,0),wᵢ) for the banded portions, with the
    open-ended top band as rateₙ·MAX(C−Lₙ,0). Zero-rate bands contribute no
    term but still advance the lower bound.
    """
    terms: list[str] = []
    lower = 0
    for band in bands:
        rate = band["ratePct"] / 100
        width = band["width"]
        if width is None:
            if rate > 0:
                terms.append(f"{rate}*MAX({ref}-{lower},0)")
            break
        if rate > 0:
            terms.append(f"{rate}*MIN(MAX({ref}-{lower},0),{width})")
        lower += width
    return "=" + "+".join(terms)


def compile_payroll(contract: PayrollContract) -> CompiledTemplate:
    cfg = load_tax_config()
    bands = cfg["paye"]["bands"]
    pension_rate = cfg["payroll"]["pension"]["employeePct"] / 100
    employer_pension_rate = cfg["payroll"]["pension"]["employerPct"] / 100
    nhf_rate = cfg["payroll"]["nhf"]["employeePct"] / 100
    nsitf_rate = cfg["payroll"]["nsitf"]["employerPct"] / 100

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[NAME_COL].width = 26
    for c in (GROSS_COL, PENSION_COL, NHF_COL, CHARGEABLE_COL, PAYE_COL, NET_COL,
              EMP_PENSION_COL, NSITF_COL):
        ws.column_dimensions[c].width = 15

    ws[f"{NAME_COL}1"] = contract.client_name
    ws[f"{NAME_COL}1"].font = TITLE
    ws[f"{NAME_COL}2"] = "PAYE & Statutory Payroll Run"
    ws[f"{NAME_COL}2"].font = BOLD
    ws[f"{NAME_COL}3"] = contract.period_label

    header_row = 5
    for col, label in (
        (NAME_COL, "Employee"),
        (GROSS_COL, "Gross (annual)"),
        (PENSION_COL, "Pension 8%"),
        (NHF_COL, "NHF 2.5%"),
        (CHARGEABLE_COL, "Chargeable"),
        (PAYE_COL, "PAYE"),
        (NET_COL, "Net pay"),
        (EMP_PENSION_COL, "Empr pension 10%"),
        (NSITF_COL, "NSITF 1%"),
    ):
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = BOLD
        cell.border = TOP_BORDER

    key: dict[str, str] = {}
    first = header_row + 1
    row = first
    for i, emp in enumerate(contract.employees):
        gross_ref = f"{GROSS_COL}{row}"
        ws[f"{NAME_COL}{row}"] = emp.name
        ws[gross_ref] = emp.gross_annual
        ws[gross_ref].number_format = NAIRA_FMT

        pension = ws[f"{PENSION_COL}{row}"]
        pension.value = f"={gross_ref}*{pension_rate}" if emp.pension else 0
        pension.number_format = NAIRA_FMT

        nhf = ws[f"{NHF_COL}{row}"]
        nhf.value = f"={gross_ref}*{nhf_rate}" if emp.nhf else 0
        nhf.number_format = NAIRA_FMT

        chargeable_ref = f"{CHARGEABLE_COL}{row}"
        chg = ws[chargeable_ref]
        chg.value = f"={gross_ref}-{PENSION_COL}{row}-{NHF_COL}{row}"
        chg.number_format = NAIRA_FMT

        paye_ref = f"{PAYE_COL}{row}"
        paye = ws[paye_ref]
        paye.value = paye_formula(bands, chargeable_ref)
        paye.number_format = NAIRA_FMT
        key[f"paye_{i}"] = paye_ref

        net_ref = f"{NET_COL}{row}"
        net = ws[net_ref]
        net.value = f"={gross_ref}-{PENSION_COL}{row}-{NHF_COL}{row}-{paye_ref}"
        net.number_format = NAIRA_FMT
        key[f"net_{i}"] = net_ref

        emp_pension = ws[f"{EMP_PENSION_COL}{row}"]
        emp_pension.value = f"={gross_ref}*{employer_pension_rate}"
        emp_pension.number_format = NAIRA_FMT

        nsitf = ws[f"{NSITF_COL}{row}"]
        nsitf.value = f"={gross_ref}*{nsitf_rate}"
        nsitf.number_format = NAIRA_FMT
        row += 1
    last = row - 1

    # Totals row.
    ws[f"{NAME_COL}{row}"] = "Totals"
    ws[f"{NAME_COL}{row}"].font = BOLD
    for col, name in (
        (GROSS_COL, "total_gross"),
        (PENSION_COL, "total_pension"),
        (NHF_COL, "total_nhf"),
        (PAYE_COL, "total_paye"),
        (NET_COL, "total_net"),
        (EMP_PENSION_COL, "total_employer_pension"),
        (NSITF_COL, "total_nsitf"),
    ):
        ref = f"{col}{row}"
        cell = ws[ref]
        cell.value = f"=SUM({col}{first}:{col}{last})"
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = TOP_BORDER
        key[name] = ref

    write_footer(ws, f"{NAME_COL}{row + 2}", COMPILER_VERSION, cfg["version"])

    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
