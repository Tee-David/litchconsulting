"""General Ledger / Categorized Bookkeeping Summary compiler.

Deterministic (PRD §3): the transaction amounts are the only raw inputs; every
column total, per-category SUMIF, and the net position are Excel formulas.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.ledger import LedgerContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "ledger-1.0.0"

DATE_COL = "B"
DESC_COL = "C"
CAT_COL = "D"
IN_COL = "E"
OUT_COL = "F"
NET_COL = "G"


def _unique_categories(contract: LedgerContract) -> list[str]:
    """Stable, sorted list of the categories present."""
    return sorted({t.category for t in contract.transactions})


def compile_ledger(contract: LedgerContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[DATE_COL].width = 14
    ws.column_dimensions[DESC_COL].width = 40
    ws.column_dimensions[CAT_COL].width = 22
    for c in (IN_COL, OUT_COL, NET_COL):
        ws.column_dimensions[c].width = 16

    ws[f"{DATE_COL}1"] = contract.client_name
    ws[f"{DATE_COL}1"].font = TITLE
    ws[f"{DATE_COL}2"] = "General Ledger — Categorized Summary"
    ws[f"{DATE_COL}2"].font = BOLD
    ws[f"{DATE_COL}3"] = contract.period_label

    key: dict[str, str] = {}

    # Transaction table.
    header_row = 5
    for col, label in (
        (DATE_COL, "Date"),
        (DESC_COL, "Description"),
        (CAT_COL, "Category"),
        (IN_COL, "Money In"),
        (OUT_COL, "Money Out"),
    ):
        cell = ws[f"{col}{header_row}"]
        cell.value = label
        cell.font = BOLD
        cell.border = TOP_BORDER

    first = header_row + 1
    row = first
    for t in contract.transactions:
        ws[f"{DATE_COL}{row}"] = t.date
        ws[f"{DESC_COL}{row}"] = t.description
        ws[f"{CAT_COL}{row}"] = t.category
        amount_col = IN_COL if t.direction == "in" else OUT_COL
        cell = ws[f"{amount_col}{row}"]
        cell.value = t.amount
        cell.number_format = NAIRA_FMT
        row += 1
    last = row - 1

    # Column totals.
    total_row = row
    ws[f"{DESC_COL}{total_row}"] = "Totals"
    ws[f"{DESC_COL}{total_row}"].font = BOLD
    for col, name in ((IN_COL, "total_in"), (OUT_COL, "total_out")):
        ref = f"{col}{total_row}"
        cell = ws[ref]
        cell.value = f"=SUM({col}{first}:{col}{last})"
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = TOP_BORDER
        key[name] = ref

    # Categorized summary block (SUMIF over the transaction rows).
    sum_title_row = total_row + 3
    ws[f"{DATE_COL}{sum_title_row}"] = "Summary by category"
    ws[f"{DATE_COL}{sum_title_row}"].font = BOLD
    sum_header_row = sum_title_row + 1
    for col, label in (
        (DESC_COL, "Category"),
        (IN_COL, "Money In"),
        (OUT_COL, "Money Out"),
        (NET_COL, "Net"),
    ):
        cell = ws[f"{col}{sum_header_row}"]
        cell.value = label
        cell.font = BOLD
        cell.border = TOP_BORDER

    cat_range = f"${CAT_COL}${first}:${CAT_COL}${last}"
    in_range = f"${IN_COL}${first}:${IN_COL}${last}"
    out_range = f"${OUT_COL}${first}:${OUT_COL}${last}"
    srow = sum_header_row + 1
    for category in _unique_categories(contract):
        ws[f"{DESC_COL}{srow}"] = category
        in_cell = ws[f"{IN_COL}{srow}"]
        in_cell.value = f'=SUMIF({cat_range},{DESC_COL}{srow},{in_range})'
        in_cell.number_format = NAIRA_FMT
        out_cell = ws[f"{OUT_COL}{srow}"]
        out_cell.value = f'=SUMIF({cat_range},{DESC_COL}{srow},{out_range})'
        out_cell.number_format = NAIRA_FMT
        net_cell = ws[f"{NET_COL}{srow}"]
        net_cell.value = f"={IN_COL}{srow}-{OUT_COL}{srow}"
        net_cell.number_format = NAIRA_FMT
        srow += 1

    # Net position.
    net_row = srow + 1
    ws[f"{DESC_COL}{net_row}"] = "Net position (in − out)"
    ws[f"{DESC_COL}{net_row}"].font = BOLD
    net_ref = f"{NET_COL}{net_row}"
    net = ws[net_ref]
    net.value = f"={key['total_in']}-{key['total_out']}"
    net.font = BOLD
    net.number_format = NAIRA_FMT
    net.border = DOUBLE_TOP
    key["net_position"] = net_ref

    write_footer(ws, f"{DATE_COL}{net_row + 2}", COMPILER_VERSION, cfg["version"])

    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
