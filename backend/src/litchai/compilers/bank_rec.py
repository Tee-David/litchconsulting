"""Bank Reconciliation Statement compiler.

Deterministic (PRD §3): the two opening balances and each adjustment amount are
the raw inputs; the section subtotals, both adjusted balances, and the
reconciliation difference are Excel formulas. A non-zero difference is the
sheet correctly reporting that the reconciliation does not tie out.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.bank_rec import BankRecContract, RecItem
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "bankrec-1.0.0"

LABEL_COL = "B"
AMOUNT_COL = "C"


def compile_bank_rec(contract: BankRecContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Rec"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 46
    ws.column_dimensions[AMOUNT_COL].width = 18

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "Bank Reconciliation Statement"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.period_label

    key: dict[str, str] = {}
    row = 5

    def base_row(label: str, amount: float) -> str:
        nonlocal row
        ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = amount
        cell.number_format = NAIRA_FMT
        row += 1
        return ref

    def section(title: str, items: list[RecItem]) -> str:
        """Header + item rows + subtotal; returns the subtotal cell ref."""
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = f"    {item.label}"
            cell = ws[f"{AMOUNT_COL}{row}"]
            cell.value = item.amount
            cell.number_format = NAIRA_FMT
            row += 1
        ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = "    Subtotal"
        sub = ws[ref]
        sub.value = f"=SUM({AMOUNT_COL}{first}:{AMOUNT_COL}{row - 1})" if items else 0
        sub.number_format = NAIRA_FMT
        sub.border = TOP_BORDER
        row += 1
        return ref

    def result_row(label: str, formula: str) -> str:
        nonlocal row
        ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = formula
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = DOUBLE_TOP
        row += 2
        return ref

    # Bank side.
    bank_ref = base_row("Balance per bank statement", contract.balance_per_bank)
    deposits_ref = section("Add: Deposits in transit", contract.deposits_in_transit)
    cheques_ref = section("Less: Outstanding cheques", contract.outstanding_cheques)
    key["adjusted_bank"] = result_row(
        "Adjusted bank balance", f"={bank_ref}+{deposits_ref}-{cheques_ref}"
    )

    # Book side.
    books_ref = base_row("Balance per cash book", contract.balance_per_books)
    adds_ref = section("Add: Credits not in books", contract.add_to_books)
    less_ref = section("Less: Charges not in books", contract.less_from_books)
    key["adjusted_book"] = result_row(
        "Adjusted cash book balance", f"={books_ref}+{adds_ref}-{less_ref}"
    )

    # Reconciliation.
    key["difference"] = result_row(
        "Difference (should be 0)", f"={key['adjusted_bank']}-{key['adjusted_book']}"
    )

    write_footer(ws, f"{LABEL_COL}{row + 1}", COMPILER_VERSION, cfg["version"])

    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
