"""Statement of Affairs / Simple Balance Sheet compiler (v1.1).

Deterministic (PRD §3): line amounts are inputs; section subtotals, total assets,
total liabilities, net assets, total equity and the balance check are formulas.
The balance check (net assets − total equity) compiles to 0 when the sheet
balances; a non-zero value is the sheet correctly reporting that it does not.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.statement_of_affairs import BsLine, StatementOfAffairsContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "soa-1.0.0"

LABEL_COL = "B"
AMT_COL = "C"


def compile_statement_of_affairs(contract: StatementOfAffairsContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Affairs"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 46
    ws.column_dimensions[AMT_COL].width = 18

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "Statement of Affairs"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.as_at_label

    key: dict[str, str] = {}
    row = 5

    def section(title: str, items: list[BsLine]) -> str:
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = f"    {item.label}"
            cell = ws[f"{AMT_COL}{row}"]
            cell.value = item.amount
            cell.number_format = NAIRA_FMT
            row += 1
        ref = f"{AMT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = "    Subtotal"
        sub = ws[ref]
        sub.value = f"=SUM({AMT_COL}{first}:{AMT_COL}{row - 1})" if items else 0
        sub.number_format = NAIRA_FMT
        sub.border = TOP_BORDER
        row += 2
        return ref

    def result(label: str, formula: str) -> str:
        nonlocal row
        ref = f"{AMT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = formula
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = DOUBLE_TOP
        row += 2
        return ref

    nca = section("Non-current assets", contract.non_current_assets)
    ca = section("Current assets", contract.current_assets)
    key["total_assets"] = result("Total assets", f"={nca}+{ca}")

    cl = section("Current liabilities", contract.current_liabilities)
    ncl = section("Non-current liabilities", contract.non_current_liabilities)
    key["total_liabilities"] = result("Total liabilities", f"={cl}+{ncl}")

    key["net_assets"] = result(
        "Net assets", f"={key['total_assets']}-{key['total_liabilities']}"
    )

    eq = section("Equity / Capital", contract.equity)
    key["total_equity"] = result("Total equity", f"={eq}")

    key["balance_check"] = result(
        "Balance check (should be 0)", f"={key['net_assets']}-{key['total_equity']}"
    )

    write_footer(ws, f"{LABEL_COL}{row}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
