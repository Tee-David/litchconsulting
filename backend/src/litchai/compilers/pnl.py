"""Profit & Loss template compiler.

Hand-written and deterministic (PRD §3): every computed cell — section totals,
gross/operating/net profit — is an Excel formula. Python never writes a
calculated number as a value, so the zero-math-error guarantee rests entirely
on the recompute gate, not on this code's arithmetic.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.pnl import LineItem, PnLContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "pnl-1.0.0"

LABEL_COL = "B"
AMOUNT_COL = "C"


def compile_pnl(contract: PnLContract) -> CompiledTemplate:
    cfg = load_tax_config()
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 42
    ws.column_dimensions[AMOUNT_COL].width = 20

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "Statement of Profit or Loss"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.period_label

    key: dict[str, str] = {}
    row = 5

    def items_section(title: str, items: list[LineItem], total_label: str) -> str:
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = item.label
            cell = ws[f"{AMOUNT_COL}{row}"]
            cell.value = item.amount
            cell.number_format = NAIRA_FMT
            row += 1
        total_ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = total_label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        total = ws[total_ref]
        total.value = f"=SUM({AMOUNT_COL}{first}:{AMOUNT_COL}{row - 1})"
        total.font = BOLD
        total.number_format = NAIRA_FMT
        total.border = TOP_BORDER
        row += 2
        return total_ref

    def result_row(label: str, formula: str, *, double: bool = False) -> str:
        nonlocal row
        ref = f"{AMOUNT_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = formula
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = DOUBLE_TOP if double else TOP_BORDER
        row += 2
        return ref

    key["total_revenue"] = items_section("Revenue", contract.revenue, "Total revenue")

    if contract.cost_of_sales:
        key["total_cost_of_sales"] = items_section(
            "Cost of sales", contract.cost_of_sales, "Total cost of sales"
        )
        key["gross_profit"] = result_row(
            "Gross profit", f"={key['total_revenue']}-{key['total_cost_of_sales']}"
        )
    else:
        key["gross_profit"] = result_row("Gross profit", f"={key['total_revenue']}")

    if contract.operating_expenses:
        key["total_operating_expenses"] = items_section(
            "Operating expenses", contract.operating_expenses, "Total operating expenses"
        )
        key["operating_profit"] = result_row(
            "Operating profit", f"={key['gross_profit']}-{key['total_operating_expenses']}"
        )
    else:
        key["operating_profit"] = result_row("Operating profit", f"={key['gross_profit']}")

    if contract.other_income:
        key["total_other_income"] = items_section(
            "Other income", contract.other_income, "Total other income"
        )
        key["net_profit"] = result_row(
            "Net profit", f"={key['operating_profit']}+{key['total_other_income']}", double=True
        )
    else:
        key["net_profit"] = result_row("Net profit", f"={key['operating_profit']}", double=True)

    write_footer(ws, f"{LABEL_COL}{row + 1}", COMPILER_VERSION, cfg["version"])

    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
