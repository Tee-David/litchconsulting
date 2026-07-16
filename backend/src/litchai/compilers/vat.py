"""VAT Returns Pack compiler (v1.1).

Deterministic (PRD §3): net amounts and the standard rate are inputs (the rate
read once from the shared tax config into a labelled cell); output VAT, input
VAT and the net payable are Excel formulas referencing that rate cell — a rate
change is one config edit, and no computed VAT value is ever hardcoded. A
negative net VAT is the sheet correctly reporting a credit to carry forward.
"""
from __future__ import annotations

from openpyxl import Workbook

from litchai.compilers._common import (
    BOLD,
    DOUBLE_TOP,
    NAIRA_FMT,
    TITLE,
    TOP_BORDER,
    CompiledTemplate,
    write_footer,
)
from litchai.contracts.vat import VatLine, VatReturnContract
from litchai.taxconfig import load_tax_config

COMPILER_VERSION = "vat-1.0.0"

LABEL_COL = "B"
NET_COL = "C"
VAT_COL = "D"


def compile_vat_return(contract: VatReturnContract) -> CompiledTemplate:
    cfg = load_tax_config()
    rate = cfg["vat"]["standardRatePct"] / 100.0

    wb = Workbook()
    ws = wb.active
    ws.title = "VAT Return"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 46
    ws.column_dimensions[NET_COL].width = 18
    ws.column_dimensions[VAT_COL].width = 18

    ws[f"{LABEL_COL}1"] = contract.client_name
    ws[f"{LABEL_COL}1"].font = TITLE
    ws[f"{LABEL_COL}2"] = "VAT Returns Pack"
    ws[f"{LABEL_COL}2"].font = BOLD
    ws[f"{LABEL_COL}3"] = contract.period_label
    if contract.tin:
        ws[f"{LABEL_COL}4"] = f"TIN: {contract.tin}"

    ws[f"{NET_COL}5"] = "Net"
    ws[f"{NET_COL}5"].font = BOLD
    ws[f"{VAT_COL}5"] = "VAT"
    ws[f"{VAT_COL}5"].font = BOLD

    # The rate is a single labelled input cell; every VAT formula references it.
    ws[f"{LABEL_COL}6"] = "VAT standard rate"
    ws[f"{LABEL_COL}6"].font = BOLD
    rate_ref = f"{NET_COL}6"
    ws[rate_ref] = rate
    ws[rate_ref].number_format = "0.00%"

    key: dict[str, str] = {}
    row = 8

    def section(title: str, items: list[VatLine], *, taxed: bool) -> tuple[str, str | None]:
        """Header + net item rows + net subtotal. When ``taxed``, also writes the
        VAT-on-subtotal formula. Returns (net_subtotal_ref, vat_ref|None)."""
        nonlocal row
        ws[f"{LABEL_COL}{row}"] = title
        ws[f"{LABEL_COL}{row}"].font = BOLD
        row += 1
        first = row
        for item in items:
            ws[f"{LABEL_COL}{row}"] = f"    {item.label}"
            net = ws[f"{NET_COL}{row}"]
            net.value = item.net_amount
            net.number_format = NAIRA_FMT
            row += 1
        net_ref = f"{NET_COL}{row}"
        ws[f"{LABEL_COL}{row}"] = "    Subtotal (net)"
        sub = ws[net_ref]
        sub.value = f"=SUM({NET_COL}{first}:{NET_COL}{row - 1})" if items else 0
        sub.number_format = NAIRA_FMT
        sub.border = TOP_BORDER
        vat_ref: str | None = None
        if taxed:
            vat_ref = f"{VAT_COL}{row}"
            vat = ws[vat_ref]
            vat.value = f"={net_ref}*{rate_ref}"
            vat.number_format = NAIRA_FMT
            vat.border = TOP_BORDER
        row += 2
        return net_ref, vat_ref

    def result_row(label: str, formula: str, col: str = VAT_COL) -> str:
        nonlocal row
        ref = f"{col}{row}"
        ws[f"{LABEL_COL}{row}"] = label
        ws[f"{LABEL_COL}{row}"].font = BOLD
        cell = ws[ref]
        cell.value = formula
        cell.font = BOLD
        cell.number_format = NAIRA_FMT
        cell.border = DOUBLE_TOP
        row += 2
        return ref

    # Output tax (VAT on sales).
    std_sales_net, output_vat = section(
        "Output tax — standard-rated sales", contract.standard_rated_sales, taxed=True
    )
    zero_net, _ = section("Zero-rated sales (0%)", contract.zero_rated_sales, taxed=False)
    exempt_net, _ = section("Exempt sales", contract.exempt_sales, taxed=False)
    key["output_vat"] = result_row("Total output VAT", f"={output_vat}")
    key["total_sales_net"] = result_row(
        "Total sales (net)", f"={std_sales_net}+{zero_net}+{exempt_net}", col=NET_COL
    )

    # Input tax (VAT on purchases) — reclaimable.
    _, input_vat = section(
        "Input tax — standard-rated purchases", contract.standard_rated_purchases, taxed=True
    )
    key["input_vat"] = result_row("Total input VAT (reclaimable)", f"={input_vat}")

    # Net VAT position.
    ws[f"{LABEL_COL}{row}"] = "Less: VAT credit brought forward"
    ws[f"{LABEL_COL}{row}"].font = BOLD
    credit_ref = f"{VAT_COL}{row}"
    ws[credit_ref].value = contract.vat_credit_brought_forward
    ws[credit_ref].number_format = NAIRA_FMT
    row += 2

    key["net_vat"] = result_row(
        "Net VAT payable / (credit c/f)",
        f"={key['output_vat']}-{key['input_vat']}-{credit_ref}",
    )

    write_footer(ws, f"{LABEL_COL}{row + 1}", COMPILER_VERSION, cfg["version"])
    return CompiledTemplate(workbook=wb, key_cells=key, compiler_version=COMPILER_VERSION)
